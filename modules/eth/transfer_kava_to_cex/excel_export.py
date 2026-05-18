"""Excel-отчёт transfer_kava_to_cex: 3 листа Wallets / Tasks / Summary."""
from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.eth.transfer_kava_to_cex import database as db

PROJECT_ROOT = Path(__file__).resolve().parents[3]
RESULT_DIR = PROJECT_ROOT / "result" / "transfer_kava_to_cex"

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_STATUS_FILL = {
    db.STATUS_PENDING:   PatternFill("solid", fgColor="FFF2CC"),
    db.STATUS_SKIPPED:   PatternFill("solid", fgColor="DDDDDD"),
    db.STATUS_TX_SENT:   PatternFill("solid", fgColor="BDD7EE"),
    db.STATUS_AWAITING:  PatternFill("solid", fgColor="FFD966"),
    db.STATUS_ARRIVED:   PatternFill("solid", fgColor="C6EFCE"),
    db.STATUS_FAILED:    PatternFill("solid", fgColor="FFC7CE"),
}

_ONE_KAVA = Decimal(10) ** 18


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, col_cells in enumerate(ws.columns, 1):
        try:
            length = max((len(str(c.value)) if c.value is not None else 0)
                         for c in col_cells)
        except ValueError:
            length = 10
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(length + 2, 12), max_width)


def _write_header(ws, headers) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B2"


def _wei_to_kava(v: Any) -> float:
    try:
        return float(Decimal(str(v or "0")) / _ONE_KAVA)
    except Exception:
        return 0.0


def export_report(out_dir: Path | None = None) -> Path:
    rows = db.list_all_tasks_joined()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = out_dir or (RESULT_DIR / f"run_{ts}")
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx = out_dir / "transfer_kava_to_cex_report.xlsx"

    wb = Workbook()

    # ---- Wallets ----
    ws = wb.active
    ws.title = "Wallets"
    headers = [
        "csv_idx", "Name", "Wallet (0x)", "CEX (kava1)", "CEX (0x)",
        "amount_spec", "Status", "Sent (KAVA)", "Arrived?",
        "src_before", "src_after", "dst_before", "dst_after",
        "tx_hash", "explorer", "attempts", "error",
    ]
    _write_header(ws, headers)
    for ri, r in enumerate(rows, start=2):
        status = r.get("status") or "—"
        values = [
            r.get("csv_index"),
            r.get("account_name") or "",
            r.get("wallet_address"),
            r.get("cex_address_bech32"),
            r.get("cex_address_evm"),
            r.get("transfer_amount_spec"),
            status,
            _wei_to_kava(r.get("sent_amount_wei")),
            "yes" if status == db.STATUS_ARRIVED else "no",
            _wei_to_kava(r.get("src_balance_before_wei")),
            _wei_to_kava(r.get("src_balance_after_wei")),
            _wei_to_kava(r.get("dst_balance_before_wei")),
            _wei_to_kava(r.get("dst_balance_after_wei")),
            r.get("tx_hash"),
            r.get("explorer_link"),
            r.get("attempts"),
            r.get("error_message"),
        ]
        for ci, v in enumerate(values, 1):
            ws.cell(row=ri, column=ci, value=v)
        fill = _STATUS_FILL.get(status)
        if fill:
            ws.cell(row=ri, column=7).fill = fill
    _autosize(ws)

    # ---- Tasks (full dump) ----
    ws = wb.create_sheet("Tasks")
    if rows:
        cols = list(rows[0].keys())
        _write_header(ws, cols)
        for ri, r in enumerate(rows, start=2):
            for ci, k in enumerate(cols, 1):
                ws.cell(row=ri, column=ci, value=r.get(k))
        _autosize(ws)

    # ---- Summary ----
    ws = wb.create_sheet("Summary")
    stats = db.get_statistics()
    summary_rows = [
        ("Generated at", datetime.now().isoformat(timespec="seconds")),
        ("DB path", str(db.DB_PATH)),
        ("Wallets total", stats.get("wallets_total", 0)),
        ("Tasks total", stats.get("total", 0)),
    ]
    for k in (db.STATUS_PENDING, db.STATUS_TX_SENT, db.STATUS_AWAITING,
              db.STATUS_ARRIVED, db.STATUS_FAILED, db.STATUS_SKIPPED):
        summary_rows.append((k, stats.get(k, 0)))
    sent_total = sum(_wei_to_kava(r.get("sent_amount_wei")) for r in rows)
    arrived_total = sum(
        _wei_to_kava(int(r.get("dst_balance_after_wei") or 0) -
                      int(r.get("dst_balance_before_wei") or 0))
        for r in rows
        if r.get("status") == db.STATUS_ARRIVED
        and r.get("dst_balance_after_wei") and r.get("dst_balance_before_wei")
    )
    summary_rows.append(("KAVA sent (cumulative)", round(sent_total, 6)))
    summary_rows.append(("KAVA arrived (cumulative)", round(arrived_total, 6)))
    for i, (k, v) in enumerate(summary_rows, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40

    wb.save(xlsx)
    return xlsx


__all__ = ["export_report", "RESULT_DIR"]
