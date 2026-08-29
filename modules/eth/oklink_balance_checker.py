"""
OKLink-based multi-token balance checker for EVM networks
(used for chains where regular RPC + ERC20 list is impractical, e.g. Polygon zkEVM).

Workflow:
  1. Load wallets from data.csv
  2. Persist per-wallet tasks in the shared SQLite DB (eth_balance_tasks)
  3. Fetch ALL token balances (native + ERC20) from OKLink internal API
  4. Save results into the DB (task balance = JSON list of tokens)
  5. After processing — export to Excel (.xlsx)
"""

from __future__ import annotations

import base64
import json
import random
import sys
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests
from rich.console import Console, Group
from rich.live import Live
from rich.panel import Panel
from rich.table import Table
from rich.text import Text

project_root = Path(__file__).parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.append(str(project_root))

from config.modules.general_config import (  # noqa: E402
    DELAY_BETWEEN_ACCOUNTS,
    NUM_THREADS,
    RETRY_COUNT,
    SLEEP_BETWEEN_ACTIONS,
)
from modules.ui import ui  # noqa: E402
from modules.ui.menu_model import BACK_KEY  # noqa: E402
from modules.eth.database import (  # noqa: E402
    create_balance_tasks,
    get_pending_tasks,
    init_database,
    reset_database_for_new_run,
    update_task_status,
)
from modules.proxy_manager import get_proxy_dict, mask_proxy, parse_proxy  # noqa: E402
from modules.simple_logger import logger  # noqa: E402

console = Console()

OKLINK_BASE = "https://www.oklink.com/api/explorer/v2"
OKLINK_SITE_INFO = "9FjOikHdpRnblJCLiskTJx0SPJiOiUGZvNmIsISQVJiOi42bpdWZyJye"

# OKLink web client computes its `x-apikey` header on every call. The static
# parts and the offset constant below come straight from their JS bundle
# (.../okfe/all-block-chain/assets/vendor-*.js, class with API_KEY field).
#
# Algorithm (reverse-engineered from the bundle):
#   1. encryptApiKey: take static UUID, rotate first 8 chars to the end.
#   2. encryptTime:   (Date.now() + ROTATE_OFFSET).toString() + 3 random digits.
#   3. comb:          base64( encryptedApiKey + "|" + encryptedTime )
_OKLINK_STATIC_UUID = "a2c903cc-b31e-4547-9299-b6d07b7631ab"
_OKLINK_ROTATED = _OKLINK_STATIC_UUID[8:] + _OKLINK_STATIC_UUID[:8]
_OKLINK_TIME_OFFSET = 1111111111111  # 13 ones


def _generate_apikey() -> str:
    """Reproduce OKLink web client's apikey generator."""
    ts = str(int(time.time() * 1000) + _OKLINK_TIME_OFFSET)
    ts += "".join(str(random.randint(0, 9)) for _ in range(3))
    return base64.b64encode(f"{_OKLINK_ROTATED}|{ts}".encode()).decode()


_devid_cache = str(uuid.uuid4())
_id_group_counter = 0


def _next_id_group() -> str:
    global _id_group_counter
    _id_group_counter += 1
    return f"{int(time.time() * 1e9) % 10**19}-c-{_id_group_counter}"


def _build_headers() -> dict:
    return {
        "accept": "application/json",
        "app-type": "web",
        "content-type": "application/json",
        "devid": _devid_cache,
        "x-apikey": _generate_apikey(),
        "x-cdn": "https://static.oklink.com",
        "x-id-group": _next_id_group(),
        "x-locale": "en_US",
        "x-simulated-trading": "undefined",
        "x-site-info": OKLINK_SITE_INFO,
        "x-utc": "0",
        "x-zkdex-env": "0",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/142.0 Safari/537.36"
        ),
        "Referer": "https://www.oklink.com/",
        "Origin": "https://www.oklink.com",
    }


def _make_proxy_dict(proxy_str: Optional[str]) -> Optional[dict]:
    """Разбор прокси общим парсером из modules.proxy_manager.

    Локальная версия просто клеила `http://` спереди, поэтому строка со схемой
    (`http://user:pass@host:port`) превращалась в невалидный двойной префикс.
    Имя функции сохранено — её импортируют планировщики swap_all_*.
    """
    return get_proxy_dict(proxy_str)


def _rand_sleep(rng) -> float:
    """Pick a random delay (seconds) from a [min, max] config range."""
    try:
        lo, hi = float(rng[0]), float(rng[1])
        if hi < lo:
            lo, hi = hi, lo
        return random.uniform(lo, hi) if hi > 0 else 0.0
    except Exception:
        return 0.0


def fetch_oklink_tokens(
    wallet: str,
    oklink_chain: str,
    proxy_dict: Optional[dict] = None,
    max_retries: Optional[int] = None,
) -> dict:
    """Return {ok: bool, tokens: [...], total_usd: float, error: str|None}."""
    if max_retries is None:
        max_retries = max(1, int(RETRY_COUNT))
    addr = wallet.lower()
    page_limit = 50
    offset = 0
    all_tokens: list[dict] = []
    total_usd = 0.0
    last_error: Optional[str] = None

    while True:
        url = (
            f"{OKLINK_BASE}/{oklink_chain}/addresses/{addr}/holders/token"
            f"?t={int(time.time()*1000)}"
        )
        body = {"offset": offset, "limit": page_limit, "valuable": False}

        success = False
        for attempt in range(max_retries):
            try:
                resp = requests.post(
                    url,
                    headers=_build_headers(),
                    json=body,
                    proxies=proxy_dict,
                    timeout=20,
                )
                if resp.status_code != 200:
                    last_error = f"HTTP {resp.status_code}"
                    time.sleep(0.5 + attempt * 0.5)
                    continue
                data = resp.json()
                if data.get("code") != 0:
                    last_error = f"API code {data.get('code')}: {data.get('msg') or data.get('detailMsg')}"
                    time.sleep(0.5 + attempt * 0.5)
                    continue
                payload = data.get("data") or {}
                hits = payload.get("hits") or []
                for h in hits:
                    all_tokens.append(
                        {
                            "symbol": h.get("symbol") or "",
                            "name": h.get("coinName") or "",
                            "value": float(h.get("value") or 0.0),
                            "price": float(h.get("price") or 0.0),
                            "usd": float(h.get("usdValue") or 0.0),
                            "contract": h.get("tokenContractAddress") or "",
                            "is_risk": bool(h.get("isRiskToken")),
                        }
                    )
                total_count = int(payload.get("total") or 0)
                offset += len(hits)
                success = True
                if not hits or offset >= total_count:
                    # Sum USD across all tokens (more transparent than the
                    # OKLink "valueTotal" field which excludes some entries).
                    total_usd = sum(t["usd"] for t in all_tokens if not t["is_risk"])
                    return {
                        "ok": True,
                        "tokens": all_tokens,
                        "total_usd": total_usd,
                        "error": None,
                    }
                # SLEEP_BETWEEN_ACTIONS — pause between paginated requests
                time.sleep(_rand_sleep(SLEEP_BETWEEN_ACTIONS))
                break  # next page
            except Exception as e:
                last_error = str(e)[:120]
                time.sleep(0.5 + attempt * 0.5)
        if not success:
            return {
                "ok": False,
                "tokens": all_tokens,
                "total_usd": total_usd,
                "error": last_error or "unknown",
            }

    # unreachable
    return {"ok": True, "tokens": all_tokens, "total_usd": total_usd, "error": None}


def _check_one(
    record: dict,
    oklink_chain: str,
    fallback_proxies: list,
) -> dict:
    """Fetch tokens for a single wallet using its paired proxy.

    `record` = {wallet, proxy, name}. If the paired proxy fails, we rotate
    through `fallback_proxies` (other rows' proxies) before giving up.
    """
    wallet = record["wallet"]
    primary_proxy = record.get("proxy")
    proxy_dict = _make_proxy_dict(primary_proxy)

    res = fetch_oklink_tokens(wallet, oklink_chain, proxy_dict)

    if not res["ok"] and fallback_proxies:
        candidates = [p for p in fallback_proxies if p and p != primary_proxy]
        for p in candidates[:2]:
            # SLEEP_BETWEEN_ACTIONS — pause before retrying with a new proxy
            time.sleep(_rand_sleep(SLEEP_BETWEEN_ACTIONS))
            res = fetch_oklink_tokens(wallet, oklink_chain, _make_proxy_dict(p))
            if res["ok"]:
                break

    res["wallet"] = wallet
    res["name"] = record.get("name") or ""
    res["proxy"] = primary_proxy or ""

    # DELAY_BETWEEN_ACCOUNTS — after this thread finishes its wallet,
    # wait before the worker is reused for the next one in the queue.
    time.sleep(_rand_sleep(DELAY_BETWEEN_ACCOUNTS))
    return res


def _build_panel(
    network: str,
    total: int,
    success: int,
    failed: int,
    logs: list,
) -> Panel:
    processed = success + failed
    percent = (processed / total * 100) if total else 0
    bar_w = 40
    filled = int(bar_w * percent / 100)
    bar = "━" * filled + "─" * (bar_w - filled)
    success_rate = (success / processed * 100) if processed else 100
    bar_style = (
        "bright_green" if success_rate >= 80 else "yellow" if success_rate >= 50 else "red"
    )

    table = Table(show_header=False, box=None, padding=(0, 1))
    table.add_column("L1", style="dim")
    table.add_column("V1")
    table.add_column("L2", style="dim")
    table.add_column("V2")
    table.add_row(
        "✅ Успешно:", Text(str(success), style="bold green"),
        "❌ Ошибки:", Text(str(failed), style="bold red"),
    )
    table.add_row(
        "📊 Всего:", Text(str(total), style="bold"),
        "⚡ Потоков:", Text(str(NUM_THREADS), style="bold cyan"),
    )

    header = Text()
    header.append("\n🌐 ", style="bold")
    header.append(f"{network}\n", style="bold cyan")
    header.append("Источник: ", style="dim")
    header.append("OKLink Explorer (web API)\n", style="bold magenta")

    pbar = Text()
    pbar.append(bar[:filled], style=f"bold {bar_style}")
    pbar.append(bar[filled:], style="dim")
    pbar.append(f" {percent:.1f}%\n", style=f"bold {bar_style}")

    proc = Text()
    proc.append("Обработано: ", style="dim")
    proc.append(f"{processed}/{total}\n\n", style="bold")

    log_section = Text()
    log_section.append("📝 Последние события:\n", style="bold")
    for ts, msg, lvl in logs[-10:]:
        log_section.append(f"  {ts} ", style="dim")
        style = {"SUCCESS": "green", "ERROR": "red", "WARNING": "yellow"}.get(lvl, "white")
        log_section.append(f"{msg}\n", style=style)
    if not logs:
        log_section.append("  Ожидание...\n", style="dim")

    return Panel(
        Group(header, pbar, proc, table, Text(""), log_section),
        title="[bold bright_blue]🪙 OKLINK MULTI-TOKEN BALANCE CHECKER[/bold bright_blue]",
        subtitle="[dim]ETHmachine[/dim]",
        border_style="bright_blue",
        padding=(1, 2),
    )


def _build_records() -> list[dict]:
    """Build [{wallet, proxy, name, private_key}] records from data.csv.

    For each row we either use `wallet_address` directly or derive the address
    from `private_key`. Each wallet is paired with the proxy from THE SAME row.
    """
    from modules.data_manager import load_data

    records: list[dict] = []
    seen: set[str] = set()
    for r in load_data():
        addr = (r.get("wallet_address") or "").strip()
        pk = (r.get("private_key") or "").strip()
        if not addr and pk:
            try:
                from eth_account import Account
                pk_hex = pk if pk.startswith("0x") else f"0x{pk}"
                addr = Account.from_key(pk_hex).address
            except Exception:
                continue
        if not addr:
            continue
        key = addr.lower()
        if key in seen:
            continue
        seen.add(key)
        records.append(
            {
                "wallet": addr,
                "proxy": (r.get("proxy") or "").strip() or None,
                "name": (r.get("name") or "").strip(),
            }
        )
    return records


def _process(
    records: list[dict],
    oklink_chain: str,
    network_name: str,
) -> dict:
    """Threaded processing. `records` = [{wallet, proxy, name}]."""
    fallback_proxies = [r["proxy"] for r in records if r.get("proxy")]
    results: dict[str, dict] = {}
    total = len(records)
    success = 0
    failed = 0
    logs: list[tuple[str, str, str]] = []
    max_workers = max(1, min(int(NUM_THREADS), total))
    logs.append((time.strftime("%H:%M:%S"), f"Запуск {total} кошельков, потоков: {max_workers}", "INFO"))

    # Нераспознанный прокси = запрос с реального IP. Молчать об этом нельзя:
    # пользователь узнает о проблеме только когда OKLink его забанит.
    bad_proxies = [p for p in dict.fromkeys(fallback_proxies) if p and parse_proxy(p) is None]
    if bad_proxies:
        logger.warning(
            f"Не удалось разобрать {len(bad_proxies)} прокси "
            f"({', '.join(mask_proxy(p) for p in bad_proxies[:3])}) — эти запросы пойдут напрямую"
        )
        logs.append((
            time.strftime("%H:%M:%S"),
            f"⚠️ Не разобрано прокси: {len(bad_proxies)} — часть запросов пойдёт напрямую",
            "WARNING"
        ))

    with Live(_build_panel(network_name, total, 0, 0, logs), console=console, refresh_per_second=4) as live:
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = {
                ex.submit(_check_one, rec, oklink_chain, fallback_proxies): rec
                for rec in records
            }
            for fut in as_completed(futures):
                rec = futures[fut]
                wallet = rec["wallet"]
                short = f"{wallet[:6]}…{wallet[-4:]}"
                name_prefix = f"{rec['name']} " if rec.get("name") else ""
                try:
                    res = fut.result(timeout=120)
                except Exception as e:
                    res = {"ok": False, "tokens": [], "total_usd": 0.0, "error": str(e)[:80], "wallet": wallet}

                results[wallet] = res
                if res["ok"]:
                    success += 1
                    update_task_status(
                        wallet,
                        "oklink_tokens",
                        "completed",
                        network=network_name,
                        balance=json.dumps(res["tokens"], ensure_ascii=False),
                        balance_usdt=f"{res['total_usd']:.6f}",
                    )
                    n = len(res["tokens"])
                    n_real = sum(1 for t in res["tokens"] if not t.get("is_risk"))
                    logs.append((
                        time.strftime("%H:%M:%S"),
                        f"{name_prefix}[{short}] ✅ {n_real}/{n} токенов | ${res['total_usd']:.4f}",
                        "SUCCESS",
                    ))
                else:
                    failed += 1
                    update_task_status(
                        wallet,
                        "oklink_tokens",
                        "failed",
                        network=network_name,
                        error_message=(res.get("error") or "")[:200],
                    )
                    logs.append((time.strftime("%H:%M:%S"), f"{name_prefix}[{short}] ❌ {res.get('error')}", "ERROR"))
                live.update(_build_panel(network_name, total, success, failed, logs))

    return results


def _export_excel(results: dict, wallets: list, network_name: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    out_dir = project_root / "result"
    out_dir.mkdir(exist_ok=True)
    clean = network_name.replace("🚀 ", "").replace(" ", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"oklink_{clean}_{ts}.xlsx"

    wb = Workbook()

    # Sheet 1 — summary per wallet
    ws = wb.active
    ws.title = "Summary"
    headers = ["#", "Name", "Wallet", "Proxy", "Tokens", "Total USD", "Status", "Error"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center")

    for i, w in enumerate(wallets, 1):
        r = results.get(w)
        if r is None:
            ws.append([i, "", w, "", 0, 0.0, "skipped", ""])
            continue
        status = "OK" if r["ok"] else "FAIL"
        err = r.get("error") or ""
        proxy_disp = (r.get("proxy") or "").split("@")[-1]
        ws.append([
            i,
            r.get("name") or "",
            w,
            proxy_disp,
            len(r.get("tokens") or []),
            round(r.get("total_usd") or 0.0, 6),
            status,
            err,
        ])

    widths = [6, 18, 46, 24, 10, 14, 10, 40]
    for c, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = wd

    # Sheet 2 — flat token list
    ws2 = wb.create_sheet("Tokens")
    headers2 = [
        "#", "Wallet", "Symbol", "Name", "Amount", "Price USD",
        "Value USD", "Contract", "Risk",
    ]
    ws2.append(headers2)
    for c in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center")

    idx = 0
    for w in wallets:
        r = results.get(w)
        if not r or not r.get("ok"):
            continue
        for tok in r.get("tokens") or []:
            idx += 1
            ws2.append([
                idx,
                w,
                tok.get("symbol", ""),
                tok.get("name", ""),
                tok.get("value", 0.0),
                tok.get("price", 0.0),
                tok.get("usd", 0.0),
                tok.get("contract", ""),
                "yes" if tok.get("is_risk") else "",
            ])

    widths = [6, 46, 14, 28, 18, 14, 14, 46, 8]
    for c, wd in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(c)].width = wd

    wb.save(path)
    return path


def _print_summary(results: dict, network: str):
    ok = [r for r in results.values() if r.get("ok")]
    bad = [r for r in results.values() if not r.get("ok")]
    total_usd = sum(r.get("total_usd") or 0.0 for r in ok)
    console.print("\n" + "=" * 60)
    console.print(f"[bold cyan]📊 ИТОГИ ({network})[/bold cyan]")
    console.print("=" * 60)
    console.print(f"[green]✅ Успешно: {len(ok)}[/green]")
    console.print(f"[red]❌ Ошибки: {len(bad)}[/red]")
    console.print(f"[yellow]💰 Суммарная стоимость активов: ${total_usd:,.4f}[/yellow]")

    top = sorted(ok, key=lambda x: x.get("total_usd") or 0.0, reverse=True)
    top = [r for r in top if (r.get("total_usd") or 0.0) > 0]
    if top:
        console.print("\n[bold cyan]🏆 Топ кошельков по стоимости:[/bold cyan]")
        for i, r in enumerate(top[:10], 1):
            wallet = r["wallet"]
            console.print(
                f"  {i}. {wallet[:10]}…{wallet[-6:]} → "
                f"[green]${r['total_usd']:.4f}[/green]  "
                f"[dim]({len(r.get('tokens') or [])} tokens)[/dim]"
            )


def run_oklink_balance_check(wallets: list | None, network_name: str, oklink_chain: str):
    """Public entry point used by the balance menu when an OKLink-backed network is chosen.

    The `wallets` argument is kept for backwards compatibility but is ignored —
    we build full records (address + paired proxy + name) directly from the
    currently selected data file via `data_manager`. This guarantees:
      • addresses are derived from `private_key` when `wallet_address` is empty
      • each request uses the proxy from the SAME row as its private key
    """
    records = _build_records()
    if not records:
        console.print(
            "[red]❌ Нет кошельков для проверки.[/red] "
            "[dim]В data файле должны быть либо `wallet_address`, либо `private_key`.[/dim]"
        )
        return

    addresses = [r["wallet"] for r in records]
    record_by_addr = {r["wallet"]: r for r in records}

    # Show user what was loaded so they can verify (this is what was confusing
    # the user previously — they couldn't tell whether requests were issued).
    proxies_count = sum(1 for r in records if r.get("proxy"))
    console.print(
        f"\n[cyan]📋 Загружено {len(records)} кошельков из data-файла "
        f"(c прокси: {proxies_count}/{len(records)})[/cyan]"
    )
    for r in records[:10]:
        addr = r["wallet"]
        name = (r.get("name") or "")[:14]
        proxy_disp = (r.get("proxy") or "—").split("@")[-1]
        console.print(
            f"   • [bold]{name:<14}[/bold] {addr[:8]}…{addr[-6:]}  "
            f"[dim]via[/dim] [yellow]{proxy_disp}[/yellow]"
        )
    if len(records) > 10:
        console.print(f"   [dim]…и ещё {len(records) - 10}[/dim]")

    action = ui.menu("Действие с базой данных", [
        ("▶️ Продолжить незавершённые задачи", "continue"),
        ("🔄 Начать заново (сброс БД)", "reset"),
        (f"{ui.glyphs.arrow} Назад", BACK_KEY),
    ])
    if action in (None, BACK_KEY):
        return

    init_database()

    if action == "reset":
        deleted = reset_database_for_new_run("oklink_tokens", network_name)
        console.print(f"[dim]🗑  Удалено старых задач: {deleted}[/dim]")

    create_balance_tasks(addresses, "oklink_tokens", network_name)
    pending = get_pending_tasks("oklink_tokens", network_name)

    from modules.eth.database import get_all_results

    if not pending:
        console.print("[yellow]⚠️ Все задачи уже выполнены — экспортирую данные из БД[/yellow]")
        rows = get_all_results("oklink_tokens", network_name)
        results = {}
        for row in rows:
            try:
                tokens = json.loads(row.get("balance") or "[]")
            except Exception:
                tokens = []
            wallet = row["wallet_address"]
            rec = record_by_addr.get(wallet) or {}
            results[wallet] = {
                "ok": row.get("status") == "completed",
                "tokens": tokens,
                "total_usd": float(row.get("balance_usdt") or 0.0),
                "error": None,
                "wallet": wallet,
                "name": rec.get("name", ""),
                "proxy": rec.get("proxy", ""),
            }
        if results:
            path = _export_excel(results, addresses, network_name)
            console.print(f"[green]💾 Excel экспорт: {path}[/green]")
            _print_summary(results, network_name)
        return

    pending_addrs = {t["wallet_address"] for t in pending}
    records_to_run = [r for r in records if r["wallet"] in pending_addrs]
    console.print(f"[cyan]📋 Задач к выполнению: {len(records_to_run)}[/cyan]\n")

    new_results = _process(records_to_run, oklink_chain, network_name)

    # Merge with already-completed entries to build a full picture for export.
    rows = get_all_results("oklink_tokens", network_name)
    full: dict[str, dict] = {}
    for row in rows:
        try:
            tokens = json.loads(row.get("balance") or "[]")
        except Exception:
            tokens = []
        wallet = row["wallet_address"]
        rec = record_by_addr.get(wallet) or {}
        full[wallet] = {
            "ok": row.get("status") == "completed",
            "tokens": tokens,
            "total_usd": float(row.get("balance_usdt") or 0.0),
            "error": None,
            "wallet": wallet,
            "name": rec.get("name", ""),
            "proxy": rec.get("proxy", ""),
        }
    full.update(new_results)

    path = _export_excel(full, addresses, network_name)
    console.print(f"\n[green]💾 Excel экспорт: {path}[/green]")
    _print_summary(full, network_name)
    console.print("\n[bold green]✅ Проверка завершена![/bold green]\n")
