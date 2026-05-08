"""Excel-отчёт swap-all zkSync Era → Base USDC."""
from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.eth.swap_all_zksync_era_to_base import database as db

PROJECT_ROOT = Path(__file__).resolve().parents[3]
RESULT_DIR = PROJECT_ROOT / "result" / "swap_all_zksync_era_to_base"

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_STATUS_FILL = {
    db.STATUS_PENDING:        PatternFill("solid", fgColor="FFF2CC"),
    db.STATUS_SKIPPED:        PatternFill("solid", fgColor="DDDDDD"),
    db.STATUS_SWAP_CREATED:   PatternFill("solid", fgColor="BDD7EE"),
    db.STATUS_TX_SENT:        PatternFill("solid", fgColor="9DC3E6"),
    db.STATUS_AWAITING:       PatternFill("solid", fgColor="FFD966"),
    db.STATUS_ARRIVED:        PatternFill("solid", fgColor="C6EFCE"),
    db.STATUS_FAILED:         PatternFill("solid", fgColor="FFC7CE"),
}


def _autosize(ws, max_width: int = 50) -> None:
    for col_idx, col_cells in enumerate(ws.columns, 1):
        try:
            length = max((len(str(c.value)) if c.value is not None else 0)
                         for c in col_cells)
        except ValueError:
            length = 10
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(length + 2, 10), max_width)


def _write_header(ws, headers: List[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B2"


def _to_decimal(v: Any) -> Decimal:
    try:
        return Decimal(str(v or "0"))
    except Exception:
        return Decimal("0")


def export_report(out_dir: Path | None = None) -> Path:
    tasks = db.list_all_tasks()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = out_dir or (RESULT_DIR / f"run_{timestamp}")
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx = out_dir / "swap_all_report.xlsx"

    wb = Workbook()

    # ---- Matrix ----
    ws = wb.active
    ws.title = "Matrix"
    wallets: List[str] = []
    seen = set()
    token_set: set[str] = set()
    by_wallet: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for t in tasks:
        w = t["wallet_address"]
        if w not in seen:
            seen.add(w)
            wallets.append(w)
        tok = t["token"] or "UNKNOWN"
        token_set.add(tok)
        by_wallet.setdefault(w, {})[tok] = t

    tokens_sorted = sorted(token_set, key=lambda x: (x not in ("USDC", "USDT"), x))
    headers = ["Wallet", "Name", "USD total", "Statuses"] + tokens_sorted
    _write_header(ws, headers)

    for row_i, w in enumerate(wallets, start=2):
        row_tokens = by_wallet.get(w, {})
        name = next((t.get("account_name") for t in row_tokens.values()
                     if t.get("account_name")), "")
        usd_total = sum(_to_decimal(t.get("usd_value")) for t in row_tokens.values())
        statuses = ",".join(sorted({t["status"] for t in row_tokens.values()}))
        ws.cell(row=row_i, column=1, value=w)
        ws.cell(row=row_i, column=2, value=name)
        ws.cell(row=row_i, column=3, value=float(usd_total))
        ws.cell(row=row_i, column=4, value=statuses)
        for col_i, tok in enumerate(tokens_sorted, start=5):
            t = row_tokens.get(tok)
            if not t:
                continue
            human = t.get("human_balance") or "0"
            try:
                value = float(human)
            except Exception:
                value = human
            cell = ws.cell(row=row_i, column=col_i, value=value)
            fill = _STATUS_FILL.get(t["status"])
            if fill:
                cell.fill = fill
    _autosize(ws)

    # ---- Tasks ----
    ws = wb.create_sheet("Tasks")
    headers = [
        "Wallet", "Name", "Token", "Contract", "Status", "Supported",
        "Human balance", "USD value", "Sent (human)", "Quote ID",
        "Bridge contract", "Src tx", "Dst tx", "Received",
        "Dst before", "Dst after", "Attempts", "Error",
    ]
    _write_header(ws, headers)
    for row_i, t in enumerate(tasks, start=2):
        values = [
            t["wallet_address"], t.get("account_name") or "",
            t.get("token"), t.get("contract"), t.get("status"),
            "yes" if t.get("supported") else "no",
            t.get("human_balance"), t.get("usd_value"),
            t.get("sent_amount_human"), t.get("swap_id"),
            t.get("deposit_address"), t.get("src_tx_hash"),
            t.get("dst_tx_hash"), t.get("received_amount"),
            t.get("dst_balance_before"), t.get("dst_balance_after"),
            t.get("attempts"), t.get("error_message"),
        ]
        for col_i, val in enumerate(values, 1):
            ws.cell(row=row_i, column=col_i, value=val)
        fill = _STATUS_FILL.get(t["status"])
        if fill:
            ws.cell(row=row_i, column=5).fill = fill
    _autosize(ws)

    # ---- Summary ----
    ws = wb.create_sheet("Summary")
    stats = db.get_statistics()
    rows = [
        ("Generated at", datetime.now().isoformat(timespec="seconds")),
        ("Wallets total", len(wallets)),
        ("Tokens total (unique)", len(tokens_sorted)),
        ("Tasks total", stats.get("total", 0)),
    ]
    for k in (db.STATUS_PENDING, db.STATUS_SKIPPED, db.STATUS_SWAP_CREATED,
              db.STATUS_TX_SENT, db.STATUS_AWAITING, db.STATUS_ARRIVED,
              db.STATUS_FAILED):
        rows.append((k, stats.get(k, 0)))
    usd_arrived = sum(_to_decimal(t.get("usd_value"))
                      for t in tasks if t["status"] == db.STATUS_ARRIVED)
    usd_pending = sum(_to_decimal(t.get("usd_value"))
                      for t in tasks
                      if t["status"] in (db.STATUS_PENDING,
                                         db.STATUS_SWAP_CREATED,
                                         db.STATUS_TX_SENT,
                                         db.STATUS_AWAITING))
    rows.append(("USD arrived", float(usd_arrived)))
    rows.append(("USD in flight / pending", float(usd_pending)))
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28

    wb.save(xlsx)
    return xlsx


__all__ = ["export_report", "RESULT_DIR"]
