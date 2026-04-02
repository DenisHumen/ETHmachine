"""Экспорт результатов xStocks в XLSX."""
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from modules.xstocks import xstocks_logger as logger
from modules.xstocks import database as db

RESULT_DIR = Path(__file__).parent.parent.parent / "result" / "xstocks"

# ---------------------------------------------------------------
# Общие стили
# ---------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _auto_width(ws, col_count, row_count, max_w=50):
    """Установить автоширину колонок."""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in range(1, min(row_count + 1, 200)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)), max_w))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 3


def _write_headers(ws, headers, row=1):
    """Записать заголовки с форматированием."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


# ---------------------------------------------------------------
# Экспорт всех данных
# ---------------------------------------------------------------

def export_results(filename: str = None) -> str | None:
    """Экспортировать все данные xStocks в XLSX.

    Листы:
      1. Wallets — все кошельки со статусами
      2. Referral Codes — все реферальные коды
      3. Stats — сводная статистика
    """
    wallets = db.get_all_wallets()
    if not wallets:
        logger.log("Нет данных для экспорта", "warning")
        return None

    os.makedirs(RESULT_DIR, exist_ok=True)

    if filename is None:
        filename = f"xstocks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = str(RESULT_DIR / filename)

    try:
        wb = Workbook()

        # ─── Лист 1: Wallets ───
        ws = wb.active
        ws.title = "Wallets"

        headers = [
            "#", "Account", "Address", "EVM Reg", "SOL Connected",
            "Referral Code", "Referral Link", "Used Referral",
            "xBoost", "Referrals", "Ref Bonus", "Today Points",
            "Total GM", "Last GM", "Next GM",
            "SOL Address",
        ]
        _write_headers(ws, headers)

        for idx, w in enumerate(wallets, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx).border = _THIN_BORDER
            ws.cell(row=row, column=2, value=w.get("account_name") or f"Wallet_{idx}").border = _THIN_BORDER
            ws.cell(row=row, column=3, value=w.get("address", "")).border = _THIN_BORDER

            # EVM Reg
            evm_cell = ws.cell(row=row, column=4, value="Yes" if w.get("evm_registered") else "No")
            evm_cell.fill = _SUCCESS_FILL if w.get("evm_registered") else _FAIL_FILL
            evm_cell.border = _THIN_BORDER

            # SOL Connected
            sol_cell = ws.cell(row=row, column=5,
                               value="Yes" if w.get("sol_connected") else ("N/A" if not w.get("sol_private_key") else "No"))
            if w.get("sol_connected"):
                sol_cell.fill = _SUCCESS_FILL
            elif w.get("sol_private_key"):
                sol_cell.fill = _FAIL_FILL
            else:
                sol_cell.fill = _WARN_FILL
            sol_cell.border = _THIN_BORDER

            ws.cell(row=row, column=6, value=w.get("referral_code", "")).border = _THIN_BORDER
            ws.cell(row=row, column=7, value=w.get("referral_link", "")).border = _THIN_BORDER
            ws.cell(row=row, column=8, value=w.get("used_referral_code", "")).border = _THIN_BORDER
            ws.cell(row=row, column=9, value=w.get("xboost", "1x")).border = _THIN_BORDER
            ws.cell(row=row, column=10, value=w.get("referrals_count", 0)).border = _THIN_BORDER
            ws.cell(row=row, column=11, value=w.get("referrals_bonus", 0)).border = _THIN_BORDER
            ws.cell(row=row, column=12, value=w.get("today_points", 0)).border = _THIN_BORDER
            ws.cell(row=row, column=13, value=w.get("total_gm_count", 0)).border = _THIN_BORDER
            ws.cell(row=row, column=14, value=(w.get("last_gm_at") or "")[:16]).border = _THIN_BORDER
            ws.cell(row=row, column=15, value=(w.get("next_gm_at") or "")[:16]).border = _THIN_BORDER
            ws.cell(row=row, column=16, value=w.get("sol_address") or "").border = _THIN_BORDER

        _auto_width(ws, len(headers), len(wallets) + 1)

        # ─── Лист 2: Referral Codes ───
        ws2 = wb.create_sheet("Referral Codes")
        ref_headers = ["#", "Wallet Address", "Referral Code", "Referral Link", "Usage Count"]
        _write_headers(ws2, ref_headers)

        ref_codes = db.get_all_referral_codes()
        for idx, rc in enumerate(ref_codes, 1):
            row = idx + 1
            ws2.cell(row=row, column=1, value=idx).border = _THIN_BORDER
            ws2.cell(row=row, column=2, value=rc.get("wallet_address", "")).border = _THIN_BORDER
            ws2.cell(row=row, column=3, value=rc.get("referral_code", "")).border = _THIN_BORDER
            ws2.cell(row=row, column=4, value=rc.get("referral_link", "")).border = _THIN_BORDER
            ws2.cell(row=row, column=5, value=rc.get("usage_count", 0)).border = _THIN_BORDER

        _auto_width(ws2, len(ref_headers), len(ref_codes) + 1)

        # ─── Лист 3: Stats Summary ───
        ws3 = wb.create_sheet("Summary")
        stats = db.get_db_stats()

        summary_data = [
            ("Total Wallets", stats.get("total_wallets", 0)),
            ("EVM Registered", stats.get("evm_registered", 0)),
            ("SOL Connected", stats.get("sol_connected", 0)),
            ("Total GM Count", stats.get("total_gm", 0)),
            ("Referral Codes", stats.get("referral_codes", 0)),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        _write_headers(ws3, ["Parameter", "Value"])
        for idx, (param, val) in enumerate(summary_data, 1):
            ws3.cell(row=idx + 1, column=1, value=param).border = _THIN_BORDER
            cell = ws3.cell(row=idx + 1, column=2, value=val)
            cell.border = _THIN_BORDER
            cell.font = Font(bold=True)

        _auto_width(ws3, 2, len(summary_data) + 1)

        wb.save(filepath)
        logger.log(f"Экспорт сохранён: {filepath}", "success")
        return filepath

    except Exception as e:
        logger.log(f"Ошибка экспорта XLSX: {e}", "error")
        return None
