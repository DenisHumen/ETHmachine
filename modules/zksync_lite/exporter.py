"""Excel экспорт результатов проверки zkSync Lite."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.simple_logger import logger
from modules.zksync_lite.database import get_all_results

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RESULT_DIR = PROJECT_ROOT / "result" / "zksync_lite"


_HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="EAF1FB")


def _autosize(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
            except Exception:
                v = ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _write_summary(ws, rows: List[Dict[str, Any]]) -> None:
    headers = [
        "wallet_address", "account_name", "status",
        "is_active", "account_id", "account_type", "pubkey_hash", "nonce",
        "eth_balance", "tokens_count", "nfts_count",
        "all_tokens", "attempts", "completed_at", "error_message",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, start=2):
        balances: Dict[str, Dict[str, Any]] = r.get("balances") or {}
        tokens_repr = "; ".join(
            f"{sym}={info.get('amount', '?')}" for sym, info in balances.items()
        )
        ws.append([
            r.get("wallet_address"),
            r.get("account_name") or "",
            r.get("status"),
            "yes" if r.get("is_active") else "no",
            r.get("account_id"),
            r.get("account_type") or "",
            r.get("pubkey_hash") or "",
            r.get("nonce"),
            r.get("eth_balance") or "",
            r.get("tokens_count") or 0,
            r.get("nfts_count") or 0,
            tokens_repr,
            r.get("attempts") or 0,
            r.get("completed_at") or "",
            r.get("error_message") or "",
        ])
        if i % 2 == 0:
            for c in ws[i]:
                c.fill = _ALT_FILL
    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_tokens(ws, rows: List[Dict[str, Any]]) -> None:
    headers = ["wallet_address", "account_name", "symbol",
               "amount", "raw_amount", "decimals"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    line_idx = 2
    for r in rows:
        balances: Dict[str, Dict[str, Any]] = r.get("balances") or {}
        if not balances:
            continue
        for sym, info in balances.items():
            ws.append([
                r.get("wallet_address"),
                r.get("account_name") or "",
                sym,
                info.get("amount"),
                info.get("raw"),
                info.get("decimals"),
            ])
            if line_idx % 2 == 0:
                for c in ws[line_idx]:
                    c.fill = _ALT_FILL
            line_idx += 1
    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_totals(ws, rows: List[Dict[str, Any]]) -> None:
    """Суммарный баланс по всем кошелькам: строка на каждый токен."""
    from decimal import Decimal, InvalidOperation

    # накапливаем суммы
    totals: Dict[str, Decimal] = {}
    wallets_per_token: Dict[str, int] = {}
    total_wallets = len(rows)
    active_wallets = sum(1 for r in rows if r.get("is_active"))

    for r in rows:
        balances: Dict[str, Dict[str, Any]] = r.get("balances") or {}
        for sym, info in balances.items():
            try:
                amt = Decimal(str(info.get("amount") or "0"))
            except InvalidOperation:
                amt = Decimal(0)
            totals[sym] = totals.get(sym, Decimal(0)) + amt
            if amt > 0:
                wallets_per_token[sym] = wallets_per_token.get(sym, 0) + 1

    # заголовок статистики
    _TEAL_FILL = PatternFill("solid", fgColor="1F7A8C")
    _TEAL_FONT = Font(bold=True, color="FFFFFF", size=13)
    _STAT_FILL = PatternFill("solid", fgColor="D9EDF7")
    _BOLD = Font(bold=True)

    ws["A1"] = "📊 Итоговые балансы zkSync Lite"
    ws["A1"].font = _TEAL_FONT
    ws["A1"].fill = _TEAL_FILL
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Всего кошельков:"
    ws["B2"] = total_wallets
    ws["B2"].font = _BOLD
    ws["A3"] = "Активных (есть аккаунт):"
    ws["B3"] = active_wallets
    ws["B3"].font = _BOLD
    ws["A4"] = "Пустых:"
    ws["B4"] = total_wallets - active_wallets
    ws["B4"].font = _BOLD

    for row_idx in range(2, 5):
        ws[f"A{row_idx}"].fill = _STAT_FILL
        ws[f"B{row_idx}"].fill = _STAT_FILL

    # разделитель
    ws.row_dimensions[5].height = 8

    # таблица токенов
    hdr_row = 6
    headers = ["Токен", "Суммарный баланс", "Кол-во кошельков (>0 баланс)",
               "% от всех активных", "Доминация (%)"]
    ws.append(headers)  # строка 6
    for c in ws[hdr_row]:
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[hdr_row].height = 20

    # сортируем: ETH первый, потом по убыванию суммы
    sorted_tokens = sorted(
        totals.items(),
        key=lambda kv: (kv[0] != "ETH", -kv[1]),
    )

    _GREEN_FONT = Font(bold=True, color="1D6A2E")
    _ALT2_FILL = PatternFill("solid", fgColor="EAF1FB")

    for line_no, (sym, total_amt) in enumerate(sorted_tokens, start=1):
        row_idx = hdr_row + line_no
        wc = wallets_per_token.get(sym, 0)
        pct_active = (wc / active_wallets * 100) if active_wallets > 0 else 0
        pct_all = (wc / total_wallets * 100) if total_wallets > 0 else 0
        ws.append([
            sym,
            str(total_amt.normalize()),
            wc,
            f"{pct_active:.1f}%",
            f"{pct_all:.1f}%",
        ])
        if line_no % 2 == 0:
            for c in ws[row_idx]:
                c.fill = _ALT2_FILL
        # выделяем ETH зелёным
        if sym == "ETH":
            ws.cell(row=row_idx, column=2).font = _GREEN_FONT

    _autosize(ws)
    ws.freeze_panes = f"A{hdr_row + 1}"


def _write_nfts(ws, rows: List[Dict[str, Any]]) -> None:
    headers = ["wallet_address", "account_name", "nft_id",
               "symbol", "address", "creator_address", "content_hash"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    line_idx = 2
    for r in rows:
        nfts: Dict[str, Dict[str, Any]] = r.get("nfts") or {}
        if not nfts:
            continue
        for nft_id, n in nfts.items():
            ws.append([
                r.get("wallet_address"),
                r.get("account_name") or "",
                nft_id,
                n.get("symbol") or "",
                n.get("address") or "",
                n.get("creator_address") or "",
                n.get("content_hash") or "",
            ])
            if line_idx % 2 == 0:
                for c in ws[line_idx]:
                    c.fill = _ALT_FILL
            line_idx += 1
    ws.freeze_panes = "A2"
    _autosize(ws)


def export_results_xlsx(target_path: Optional[Path] = None) -> Optional[Path]:
    rows = get_all_results()
    if not rows:
        logger.warning("zkSync Lite: нет данных для экспорта")
        return None

    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    if target_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_path = RESULT_DIR / f"zksync_lite_{ts}.xlsx"

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "summary"
    _write_summary(ws_sum, rows)

    ws_totals = wb.create_sheet("totals")
    _write_totals(ws_totals, rows)

    ws_tokens = wb.create_sheet("tokens")
    _write_tokens(ws_tokens, rows)

    ws_nfts = wb.create_sheet("nfts")
    _write_nfts(ws_nfts, rows)

    wb.save(str(target_path))
    return target_path


__all__ = ["export_results_xlsx", "RESULT_DIR"]
