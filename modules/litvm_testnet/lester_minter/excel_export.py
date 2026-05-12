"""Excel-отчёт для Lester Minter.

Формат: 2 листа — Deployments (по одной строке на токен) и Wallets (агрегаты).
Сохраняется в `result/lester_minter/run_<ts>/lester_minter_report.xlsx`.
"""
from __future__ import annotations

import time
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.litvm_testnet.lester_minter import database as db


_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_CELL_FONT = Font(name="Calibri", size=10)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _fmt_ts(epoch: Optional[float]) -> str:
    if not epoch:
        return ""
    return datetime.fromtimestamp(float(epoch)).strftime("%Y-%m-%d %H:%M:%S")


def _format_supply_human(supply_wei: str, decimals: int) -> str:
    try:
        whole = int(supply_wei) // (10 ** int(decimals)) if int(decimals) > 0 else int(supply_wei)
    except Exception:
        return supply_wei
    return f"{whole:,}".replace(",", "_")


def _features_label(m, b, p) -> str:
    flags = []
    if int(m or 0): flags.append("Mintable")
    if int(b or 0): flags.append("Burnable")
    if int(p or 0): flags.append("Pausable")
    return ", ".join(flags) or "—"


def _style_header(ws, row: int, columns: list[str]) -> None:
    for col, title in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 10
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            ln = len(str(v))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[letter].width = min(max_len + 2, 64)


def build_report(out_dir: Optional[Path] = None) -> Path:
    deployments = db.list_all_deployments()
    wallets = db.list_all_wallets()

    if out_dir is None:
        out_dir = Path("result") / "lester_minter" / f"run_{time.strftime('%Y%m%d_%H%M%S')}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "lester_minter_report.xlsx"

    wb = Workbook()

    # Sheet 1: Deployments
    ws = wb.active
    ws.title = "Deployments"
    cols = [
        "Wallet", "TX#", "Status", "Token Name", "Symbol", "Decimals",
        "Total Supply", "Features", "Token Address", "Tx Hash",
        "Gas Used", "Fee (zkLTC)", "Logo URL", "Created", "Confirmed",
        "Error",
    ]
    _style_header(ws, 1, cols)
    for i, d in enumerate(deployments, 2):
        fee_zk = float(int(d.get("fee_wei") or 0)) / 1e18
        row = [
            d.get("address"),
            int(d.get("tx_index") or 0),
            d.get("status"),
            d.get("token_name"),
            d.get("token_symbol"),
            int(d.get("decimals") or 0),
            _format_supply_human(d.get("total_supply") or "0",
                                 int(d.get("decimals") or 0)),
            _features_label(d.get("mintable"), d.get("burnable"), d.get("pausable")),
            d.get("token_address"),
            d.get("tx_hash"),
            int(d.get("gas_used") or 0),
            f"{fee_zk:.6f}",
            d.get("logo_url"),
            _fmt_ts(d.get("created_at")),
            _fmt_ts(d.get("confirmed_at")),
            d.get("error_message"),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _CELL_FONT
            cell.alignment = _CENTER if col in (2, 3, 6, 11) else _LEFT
    ws.freeze_panes = "A2"
    _autosize(ws)

    # Sheet 2: Wallets summary
    ws2 = wb.create_sheet("Wallets")
    cols2 = ["Address", "Name", "Status", "Planned", "Completed",
             "Failed", "Created", "Updated", "Error"]
    _style_header(ws2, 1, cols2)
    for i, w in enumerate(wallets, 2):
        row = [
            w.get("address"), w.get("name"), w.get("status"),
            int(w.get("planned") or 0), int(w.get("completed") or 0),
            int(w.get("failed") or 0),
            _fmt_ts(w.get("created_at")), _fmt_ts(w.get("updated_at")),
            w.get("error_message"),
        ]
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = _CELL_FONT
            cell.alignment = _CENTER if col in (3, 4, 5, 6) else _LEFT
    ws2.freeze_panes = "A2"
    _autosize(ws2)

    wb.save(out_file)
    return out_file
