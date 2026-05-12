"""Excel-отчёт для Onmi Liquidity."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.litvm_testnet.onmi.liquidity import database as db


REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
RESULT_DIR = REPO_ROOT / "result" / "onmi_liquidity"

HEADER_FILL = PatternFill("solid", fgColor="2F4F6F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _ts_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 6
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            min(max_len + 2, max_width)
        )


def _fmt_wei(wei) -> Optional[float]:
    if wei in (None, "", "None"):
        return None
    try:
        return int(wei) / 1e18
    except Exception:
        return None


def _ts(v) -> Optional[str]:
    if not v:
        return None
    try:
        return datetime.fromtimestamp(float(v), tz=timezone.utc).strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except Exception:
        return None


def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


def build_report() -> Path:
    positions = db.list_positions()
    history = db.list_history(limit=100_000)
    stats = db.get_statistics()

    out_dir = RESULT_DIR / f"run_{_ts_now()}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "onmi_liquidity_report.xlsx"

    wb = Workbook()

    # Positions
    ws = wb.active
    ws.title = "Positions"
    ph = [
        "wallet_address", "token_symbol", "token_address", "pair_address",
        "lp_net", "lp_acquired", "lp_removed",
        "total_added_eth", "total_added_token",
        "total_removed_eth", "total_removed_token",
        "first_action_at", "last_action_at",
    ]
    _write_header(ws, ph)
    for row, p in enumerate(positions, 2):
        ws.cell(row=row, column=1, value=p.get("wallet_address"))
        ws.cell(row=row, column=2, value=p.get("token_symbol"))
        ws.cell(row=row, column=3, value=p.get("token_address"))
        ws.cell(row=row, column=4, value=p.get("pair_address"))
        ws.cell(row=row, column=5, value=_fmt_wei(p.get("lp_net_wei")))
        ws.cell(row=row, column=6, value=_fmt_wei(p.get("lp_acquired_wei")))
        ws.cell(row=row, column=7, value=_fmt_wei(p.get("lp_removed_wei")))
        ws.cell(row=row, column=8, value=_fmt_wei(p.get("total_added_eth_wei")))
        ws.cell(row=row, column=9, value=_fmt_wei(p.get("total_added_token_wei")))
        ws.cell(row=row, column=10, value=_fmt_wei(p.get("total_removed_eth_wei")))
        ws.cell(row=row, column=11, value=_fmt_wei(p.get("total_removed_token_wei")))
        ws.cell(row=row, column=12, value=_ts(p.get("first_action_at")))
        ws.cell(row=row, column=13, value=_ts(p.get("last_action_at")))
    _autosize(ws)

    # History
    ws2 = wb.create_sheet("History")
    hh = [
        "id", "wallet_name", "wallet_address", "side", "token_symbol",
        "token_address", "pair_address", "amount_eth", "amount_token",
        "lp_tokens", "tx_hash", "gas_used", "status", "attempts",
        "error", "created_at", "sent_at", "confirmed_at",
    ]
    _write_header(ws2, hh)
    for row, t in enumerate(history, 2):
        ws2.cell(row=row, column=1, value=t.get("id"))
        ws2.cell(row=row, column=2, value=t.get("wallet_name"))
        ws2.cell(row=row, column=3, value=t.get("wallet_address"))
        ws2.cell(row=row, column=4, value=t.get("side"))
        ws2.cell(row=row, column=5, value=t.get("token_symbol"))
        ws2.cell(row=row, column=6, value=t.get("token_address"))
        ws2.cell(row=row, column=7, value=t.get("pair_address"))
        ws2.cell(row=row, column=8, value=_fmt_wei(t.get("amount_eth_wei")))
        ws2.cell(row=row, column=9, value=_fmt_wei(t.get("amount_token_wei")))
        ws2.cell(row=row, column=10, value=_fmt_wei(t.get("lp_tokens_wei")))
        ws2.cell(row=row, column=11, value=t.get("tx_hash"))
        ws2.cell(row=row, column=12, value=t.get("gas_used"))
        ws2.cell(row=row, column=13, value=t.get("status"))
        ws2.cell(row=row, column=14, value=t.get("attempts"))
        ws2.cell(row=row, column=15, value=t.get("error_message"))
        ws2.cell(row=row, column=16, value=_ts(t.get("created_at")))
        ws2.cell(row=row, column=17, value=_ts(t.get("sent_at")))
        ws2.cell(row=row, column=18, value=_ts(t.get("confirmed_at")))
    _autosize(ws2)

    # Summary
    ws3 = wb.create_sheet("Summary")
    _write_header(ws3, ["metric", "value"])
    for i, (k, v) in enumerate(stats.items(), 2):
        ws3.cell(row=i, column=1, value=k)
        ws3.cell(row=i, column=2, value=v)
    _autosize(ws3)

    wb.save(out_path)
    return out_path
