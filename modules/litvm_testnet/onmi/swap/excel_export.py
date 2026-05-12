"""Excel-отчёт для модуля Onmi Swap."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.litvm_testnet.onmi.swap import database as db


REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
RESULT_DIR = REPO_ROOT / "result" / "onmi_swap"

HEADER_FILL = PatternFill("solid", fgColor="2F4F6F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _ts_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 6
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            min(max_len + 2, max_width)
        )


def _fmt_wei(wei) -> Optional[float]:
    if wei in (None, "", "None"):
        return None
    try:
        return int(wei) / 1e18
    except Exception:
        return None


def _ts(v) -> Optional[str]:
    if not v:
        return None
    try:
        return datetime.fromtimestamp(float(v), tz=timezone.utc).strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except Exception:
        return None


def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


def build_report() -> Path:
    swaps = db.list_swaps(limit=100_000)
    pairs = db.list_known_pairs(only_enabled=False)
    stats = db.get_statistics()

    out_dir = RESULT_DIR / f"run_{_ts_now()}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "onmi_swap_report.xlsx"

    wb = Workbook()

    # Swaps
    ws = wb.active
    ws.title = "Swaps"
    headers = [
        "id", "wallet_name", "wallet_address", "side", "token_symbol",
        "token_address", "pair_address", "amount_in", "amount_out",
        "min_out", "tx_hash", "gas_used", "status", "attempts", "error",
        "created_at", "sent_at", "confirmed_at",
    ]
    _write_header(ws, headers)
    for row, t in enumerate(swaps, 2):
        ws.cell(row=row, column=1, value=t.get("id"))
        ws.cell(row=row, column=2, value=t.get("wallet_name"))
        ws.cell(row=row, column=3, value=t.get("wallet_address"))
        ws.cell(row=row, column=4, value=t.get("side"))
        ws.cell(row=row, column=5, value=t.get("token_symbol"))
        ws.cell(row=row, column=6, value=t.get("token_address"))
        ws.cell(row=row, column=7, value=t.get("pair_address"))
        ws.cell(row=row, column=8, value=_fmt_wei(t.get("amount_in_wei")))
        ws.cell(row=row, column=9, value=_fmt_wei(t.get("amount_out_wei")))
        ws.cell(row=row, column=10, value=_fmt_wei(t.get("min_out_wei")))
        ws.cell(row=row, column=11, value=t.get("tx_hash"))
        ws.cell(row=row, column=12, value=t.get("gas_used"))
        ws.cell(row=row, column=13, value=t.get("status"))
        ws.cell(row=row, column=14, value=t.get("attempts"))
        ws.cell(row=row, column=15, value=t.get("error_message"))
        ws.cell(row=row, column=16, value=_ts(t.get("created_at")))
        ws.cell(row=row, column=17, value=_ts(t.get("sent_at")))
        ws.cell(row=row, column=18, value=_ts(t.get("confirmed_at")))
    _autosize(ws)

    # Pairs
    ws2 = wb.create_sheet("KnownPairs")
    ph = ["pair_address", "token_symbol", "token_address", "token_decimals",
          "reserve_native_zkLTC", "reserve_token", "disabled",
          "discovered_at", "last_seen_at"]
    _write_header(ws2, ph)
    for row, p in enumerate(pairs, 2):
        ws2.cell(row=row, column=1, value=p.get("pair_address"))
        ws2.cell(row=row, column=2, value=p.get("token_symbol"))
        ws2.cell(row=row, column=3, value=p.get("token_address"))
        ws2.cell(row=row, column=4, value=p.get("token_decimals"))
        ws2.cell(row=row, column=5, value=_fmt_wei(p.get("reserve_native_wei")))
        ws2.cell(row=row, column=6, value=_fmt_wei(p.get("reserve_token_wei")))
        ws2.cell(row=row, column=7, value=p.get("disabled"))
        ws2.cell(row=row, column=8, value=_ts(p.get("discovered_at")))
        ws2.cell(row=row, column=9, value=_ts(p.get("last_seen_at")))
    _autosize(ws2)

    # Summary
    ws3 = wb.create_sheet("Summary")
    _write_header(ws3, ["metric", "value"])
    for i, (k, v) in enumerate(stats.items(), 2):
        ws3.cell(row=i, column=1, value=k)
        ws3.cell(row=i, column=2, value=v)
    _autosize(ws3)

    wb.save(out_path)
    return out_path
