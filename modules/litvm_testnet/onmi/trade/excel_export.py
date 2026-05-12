"""Excel-отчёт для модуля Onmi Trade.

Листы:
  • Trades — каждая buy/sell-операция (фильтр для статистики удобно делать
    в Excel).
  • KnownTokens — текущий список известных токенов (для понимания пула).
  • Summary — агрегаты по кошелькам и по токенам.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.litvm_testnet.onmi.trade import database as db


REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
RESULT_DIR = REPO_ROOT / "result" / "onmi_trade"


def _ts_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 6
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            min(max_len + 2, max_width)
        )


def _fmt_wei(wei) -> Optional[float]:
    if wei in (None, "", "None"):
        return None
    try:
        return int(wei) / 1e18
    except Exception:
        return None


def _ts(v) -> Optional[str]:
    if not v:
        return None
    try:
        return datetime.fromtimestamp(float(v), tz=timezone.utc).strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except Exception:
        return None


HEADER_FILL = PatternFill("solid", fgColor="2F4F6F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


def build_report() -> Path:
    trades = db.list_trades(limit=100_000)
    tokens = db.list_known_tokens(include_graduated=True)
    stats = db.get_statistics()

    out_dir = RESULT_DIR / f"run_{_ts_now()}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "onmi_trade_report.xlsx"

    wb = Workbook()

    # --- Sheet 1: Trades --------------------------------------------------
    ws = wb.active
    ws.title = "Trades"
    headers = [
        "id", "wallet_name", "wallet_address", "token_symbol",
        "token_address", "side", "amount_in", "amount_out", "tx_hash",
        "gas_used", "status", "attempts", "error",
        "created_at", "sent_at", "confirmed_at",
    ]
    _write_header(ws, headers)
    for row, t in enumerate(trades, 2):
        ws.cell(row=row, column=1, value=t.get("id"))
        ws.cell(row=row, column=2, value=t.get("wallet_name"))
        ws.cell(row=row, column=3, value=t.get("wallet_address"))
        ws.cell(row=row, column=4, value=t.get("token_symbol"))
        ws.cell(row=row, column=5, value=t.get("token_address"))
        ws.cell(row=row, column=6, value=t.get("side"))
        ws.cell(row=row, column=7, value=_fmt_wei(t.get("amount_in_wei")))
        ws.cell(row=row, column=8, value=_fmt_wei(t.get("amount_out_wei")))
        ws.cell(row=row, column=9, value=t.get("tx_hash"))
        ws.cell(row=row, column=10, value=t.get("gas_used"))
        ws.cell(row=row, column=11, value=t.get("status"))
        ws.cell(row=row, column=12, value=t.get("attempts"))
        ws.cell(row=row, column=13, value=(t.get("error_message") or "")[:300])
        ws.cell(row=row, column=14, value=_ts(t.get("created_at")))
        ws.cell(row=row, column=15, value=_ts(t.get("sent_at")))
        ws.cell(row=row, column=16, value=_ts(t.get("confirmed_at")))
    _autosize(ws)

    # --- Sheet 2: KnownTokens ---------------------------------------------
    ws2 = wb.create_sheet("KnownTokens")
    headers2 = [
        "address", "symbol", "name", "creator_address", "source",
        "graduated", "created_at", "last_seen_at",
    ]
    _write_header(ws2, headers2)
    for row, t in enumerate(tokens, 2):
        ws2.cell(row=row, column=1, value=t.get("address"))
        ws2.cell(row=row, column=2, value=t.get("symbol"))
        ws2.cell(row=row, column=3, value=t.get("name"))
        ws2.cell(row=row, column=4, value=t.get("creator_address"))
        ws2.cell(row=row, column=5, value=t.get("source"))
        ws2.cell(row=row, column=6, value=bool(t.get("graduated")))
        ws2.cell(row=row, column=7, value=_ts(t.get("created_at")))
        ws2.cell(row=row, column=8, value=_ts(t.get("last_seen_at")))
    _autosize(ws2)

    # --- Sheet 3: Summary --------------------------------------------------
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Метрика"; ws3["B1"] = "Значение"
    for c in ws3["A1:B1"][0]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    keys = [
        ("known_tokens", "Known tokens"),
        ("graduated", "Graduated"),
        ("trades_total", "Trades total"),
        ("status_arrived", "  · arrived"),
        ("status_failed", "  · failed"),
        ("status_sent", "  · sent (in-flight)"),
        ("status_pending", "  · pending"),
        ("side_buy", "Successful buys"),
        ("side_sell", "Successful sells"),
    ]
    row = 2
    for k, label in keys:
        ws3.cell(row=row, column=1, value=label)
        ws3.cell(row=row, column=2, value=stats.get(k, 0))
        row += 1

    # per-wallet breakdown
    row += 1
    ws3.cell(row=row, column=1, value="Per-wallet breakdown:").font = Font(bold=True)
    row += 1
    headers3 = ["wallet", "buys", "sells", "zkLTC spent", "zkLTC received", "failures"]
    for i, h in enumerate(headers3):
        c = ws3.cell(row=row, column=i + 1, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    row += 1

    per_wallet: dict[str, dict] = {}
    for t in trades:
        w = (t.get("wallet_name") or t.get("wallet_address") or "?")
        d = per_wallet.setdefault(w, {
            "buys": 0, "sells": 0, "spent": 0.0, "received": 0.0, "failures": 0,
        })
        st = (t.get("status") or "").lower()
        side = (t.get("side") or "").lower()
        if st == "arrived":
            if side == "buy":
                d["buys"] += 1
                d["spent"] += (_fmt_wei(t.get("amount_in_wei")) or 0.0)
            elif side == "sell":
                d["sells"] += 1
                d["received"] += (_fmt_wei(t.get("amount_out_wei")) or 0.0)
        elif st == "failed":
            d["failures"] += 1

    for w, d in sorted(per_wallet.items(), key=lambda x: -(x[1]["buys"] + x[1]["sells"])):
        ws3.cell(row=row, column=1, value=w)
        ws3.cell(row=row, column=2, value=d["buys"])
        ws3.cell(row=row, column=3, value=d["sells"])
        ws3.cell(row=row, column=4, value=round(d["spent"], 8))
        ws3.cell(row=row, column=5, value=round(d["received"], 8))
        ws3.cell(row=row, column=6, value=d["failures"])
        row += 1

    _autosize(ws3)

    wb.save(out_path)
    return out_path
