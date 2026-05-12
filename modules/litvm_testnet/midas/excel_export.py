"""Excel-отчёт для Midas Prediction Market.

Формат: 4 листа — Bets, Wallets, FaucetClaims, CheckIns.
Сохраняется в `result/midas/run_<ts>/midas_report.xlsx`.
"""
from __future__ import annotations

import time
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.litvm_testnet.midas import database as db


_HEADER_FILL = PatternFill(start_color="305496", end_color="305496",
                           fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_CELL_FONT = Font(name="Calibri", size=10)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _fmt_ts(epoch: Optional[float]) -> str:
    if not epoch:
        return ""
    try:
        return datetime.fromtimestamp(float(epoch)).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def _style_header(ws, row: int, columns: list[str]) -> None:
    for col, title in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 10
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            ln = len(str(v))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[letter].width = min(max_len + 2, 64)


def build_report(out_dir: Optional[Path] = None) -> Path:
    bets = db.list_all_bets()
    wallets = db.list_all_wallets()
    faucets = db.list_all_faucet_claims()
    checkins = db.list_all_checkins()

    if out_dir is None:
        out_dir = Path("result") / "midas" / f"run_{time.strftime('%Y%m%d_%H%M%S')}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "midas_report.xlsx"

    wb = Workbook()

    # 1. Bets
    ws = wb.active
    ws.title = "Bets"
    cols = ["Wallet", "Market", "Title", "Outcome", "Amount USDC",
            "Shares", "Max Cost (raw)", "Status",
            "Approve Tx", "Buy Tx", "Gas Used",
            "Created", "Sent", "Confirmed", "Error"]
    _style_header(ws, 1, cols)
    for i, b in enumerate(bets, 2):
        row = [
            b.get("address"),
            b.get("market_address"),
            b.get("market_title"),
            int(b.get("outcome_index") or 0),
            float(b.get("amount_usdc_human") or 0.0),
            b.get("shares"),
            b.get("max_cost_raw"),
            b.get("status"),
            b.get("approve_tx_hash"),
            b.get("buy_tx_hash"),
            int(b.get("gas_used") or 0),
            _fmt_ts(b.get("created_at")),
            _fmt_ts(b.get("sent_at")),
            _fmt_ts(b.get("confirmed_at")),
            b.get("error_message"),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = _CELL_FONT
            cell.alignment = _CENTER if col in (4, 5, 8, 11) else _LEFT
    ws.freeze_panes = "A2"
    _autosize(ws)

    # 2. Wallets
    ws2 = wb.create_sheet("Wallets")
    cols2 = ["Address", "Name", "Nickname", "Registered", "Status",
             "Bets Planned", "Bets Done", "Bets Failed",
             "Last USDC Faucet", "Last zkLTC Faucet", "Last Check-in",
             "Created", "Updated", "Error"]
    _style_header(ws2, 1, cols2)
    for i, w in enumerate(wallets, 2):
        row = [
            w.get("address"), w.get("name"), w.get("nickname"),
            int(w.get("registered") or 0), w.get("status"),
            int(w.get("bets_planned") or 0),
            int(w.get("bets_completed") or 0),
            int(w.get("bets_failed") or 0),
            _fmt_ts(w.get("last_usdc_faucet_at")),
            _fmt_ts(w.get("last_native_faucet_at")),
            _fmt_ts(w.get("last_checkin_at")),
            _fmt_ts(w.get("created_at")),
            _fmt_ts(w.get("updated_at")),
            w.get("error_message"),
        ]
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = _CELL_FONT
            cell.alignment = _CENTER if col in (4, 5, 6, 7, 8) else _LEFT
    ws2.freeze_panes = "A2"
    _autosize(ws2)

    # 3. FaucetClaims
    ws3 = wb.create_sheet("FaucetClaims")
    cols3 = ["ID", "Wallet", "Kind", "Status", "Created",
             "Response", "Error"]
    _style_header(ws3, 1, cols3)
    for i, f in enumerate(faucets, 2):
        row = [
            int(f.get("id") or 0), f.get("address"), f.get("kind"),
            f.get("status"), _fmt_ts(f.get("created_at")),
            f.get("response"), f.get("error_message"),
        ]
        for col, val in enumerate(row, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.font = _CELL_FONT
            cell.alignment = _CENTER if col in (1, 3, 4) else _LEFT
    ws3.freeze_panes = "A2"
    _autosize(ws3)

    # 4. CheckIns
    ws4 = wb.create_sheet("CheckIns")
    cols4 = ["ID", "Wallet", "Status", "Streak", "Created",
             "Response", "Error"]
    _style_header(ws4, 1, cols4)
    for i, c in enumerate(checkins, 2):
        row = [
            int(c.get("id") or 0), c.get("address"),
            c.get("status"),
            int(c.get("streak") or 0) if c.get("streak") is not None else None,
            _fmt_ts(c.get("created_at")),
            c.get("response"), c.get("error_message"),
        ]
        for col, val in enumerate(row, 1):
            cell = ws4.cell(row=i, column=col, value=val)
            cell.font = _CELL_FONT
            cell.alignment = _CENTER if col in (1, 3, 4) else _LEFT
    ws4.freeze_panes = "A2"
    _autosize(ws4)

    wb.save(out_file)
    return out_file
