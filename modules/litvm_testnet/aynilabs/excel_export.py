"""Excel-отчёт для модуля Aynilabs.

3 листа (AGENTS.md §6): Matrix · Tasks · Summary.
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.litvm_testnet.aynilabs import database as db


REPO_ROOT = Path(__file__).resolve().parent.parent.parent.parent
RESULT_DIR = REPO_ROOT / "result" / "aynilabs"


def _ts_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 6
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            min(max_len + 2, max_width)
        )


def _fmt_wei(wei: Optional[str | int]) -> Optional[float]:
    if wei in (None, "", "None"):
        return None
    try:
        return int(wei) / 1e18
    except Exception:  # noqa: BLE001
        return None


def _ts(v: Optional[float]) -> Optional[str]:
    if not v:
        return None
    return datetime.fromtimestamp(float(v), tz=timezone.utc).strftime(
        "%Y-%m-%d %H:%M:%S"
    )


def build_report() -> Path:
    tasks = db.list_all_tasks()
    stats = db.get_statistics()

    out_dir = RESULT_DIR / f"run_{_ts_now()}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "aynilabs_report.xlsx"

    wb = Workbook()

    # --- Matrix --------------------------------------------------------------
    ws_m = wb.active
    ws_m.title = "Matrix"
    headers = ["#", "Name", "Wallet", "Status",
               "Planned zkLTC", "Received WzkLTC", "TX hash"]
    ws_m.append(headers)
    for c in ws_m[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    for i, t in enumerate(tasks, 1):
        ws_m.append([
            i,
            t.get("name") or "",
            t["address"],
            t["status"],
            _fmt_wei(t.get("planned_amount_wei")),
            _fmt_wei(t.get("received_amount_wei")),
            t.get("tx_hash") or "",
        ])
    _autosize(ws_m)

    # --- Tasks ---------------------------------------------------------------
    ws_t = wb.create_sheet("Tasks")
    cols = [
        "id", "address", "name", "tx_index", "status",
        "planned_amount_wei", "planned_amount_human",
        "native_balance_before_wei",
        "wzkltc_balance_before_wei", "wzkltc_balance_after_wei",
        "received_amount_wei", "tx_hash", "gas_used",
        "attempts", "error_message",
        "created_at", "sent_at", "confirmed_at", "updated_at",
    ]
    ws_t.append(cols)
    for c in ws_t[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    for t in tasks:
        ws_t.append([
            t.get(c) if c not in ("created_at", "sent_at", "confirmed_at",
                                  "updated_at")
            else _ts(t.get(c))
            for c in cols
        ])
    _autosize(ws_t)

    # --- Summary -------------------------------------------------------------
    ws_s = wb.create_sheet("Summary")
    ws_s.append(["Aynilabs wrap report"])
    ws_s["A1"].font = Font(bold=True, size=14)
    ws_s.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws_s.append([])
    ws_s.append(["Status", "Count"])
    for c in ws_s[4]:
        c.font = Font(bold=True)
    for key in ("pending", "tx_sent", "arrived", "failed", "skipped"):
        ws_s.append([key, stats.get(key, 0)])
    ws_s.append(["TOTAL", stats.get("total", 0)])
    ws_s[ws_s.max_row][0].font = Font(bold=True)
    ws_s[ws_s.max_row][1].font = Font(bold=True)
    # totals по deposit'ам
    total_planned = sum((int(t.get("planned_amount_wei") or 0) for t in tasks))
    total_received = sum(
        (int(t.get("received_amount_wei") or 0) for t in tasks)
    )
    ws_s.append([])
    ws_s.append(["Planned zkLTC total", total_planned / 1e18])
    ws_s.append(["Received WzkLTC total", total_received / 1e18])
    _autosize(ws_s)

    wb.save(out_path)
    return out_path
