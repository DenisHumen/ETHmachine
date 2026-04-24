"""
Base Network Analytics — Dune wallet checker (БЕЗ API-ключа).

Дашборд: https://dune.com/nvthao/base-network-analytics-dashboard

Чекер вместо платного Dune API использует публичный дашборд: через
patchright-Chromium (обход Cloudflare) открывает страницу, вводит адрес
в строку поиска каждой таблицы и парсит найденную строку. Окно браузера
позиционируется offscreen, поэтому визуально ничего не появляется.

Адреса берутся из выбранного data/data*.csv:
  • `private_key` → EVM-адрес через `eth_account`;
  • если приватного ключа нет — используется `wallet_address`;
  • у каждого кошелька свой `proxy` из той же строки CSV
    (прокси пробрасывается в browser-context).

Состояние и результаты хранятся в SQLite `db/dune_base_checker.db`,
что позволяет продолжить с места остановки, повторно выгрузить
результаты в Excel без повторного парсинга или очистить и начать заново.

Используются:
  • config/modules/cfg_base.py — потоки, задержки, RETRY_COUNT
  • modules/data_manager.py    — загрузка строк data.csv
  • modules/proxy_manager.py   — маскирование прокси в логах
  • modules/simple_logger.py   — единый формат логирования
  • modules/dune/base/scraper.py — Playwright-скрейпер дашборда
"""

from __future__ import annotations

import json
import random
import time
from datetime import datetime
from pathlib import Path
from queue import Empty, Queue
from threading import Event, Lock, Thread
from typing import Any, Dict, List, Optional, Tuple

from colorama import Fore, Style
from eth_account import Account
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tqdm import tqdm

from config.modules.cfg_base import (
    DELAY_BETWEEN_ACCOUNTS,
    NUM_THREADS,
    RETRY_COUNT,
    SLEEP_BETWEEN_ACTIONS,
)
from modules.data_manager import load_data
from modules.dune.base.database import (
    create_check_tasks,
    get_all_results,
    get_pending_addresses,
    get_task_statistics,
    init_database,
    update_task_failed,
    update_task_success,
)
from modules.dune.base.scraper import DuneBaseScraper
from modules.proxy_manager import mask_proxy
from modules.simple_logger import log_task, log_wallet_task, logger, tqdm_safe_logging

# ──────────────────────────────────────────────────────────────────────────
# Константы
# ──────────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parents[3]
RESULT_DIR = PROJECT_ROOT / "result" / "dune"

_stop_event = Event()

# ──────────────────────────────────────────────────────────────────────────
# Прогресс-бар (pip-style), обновляется из worker-потоков
# ──────────────────────────────────────────────────────────────────────────

_pbar: Optional[tqdm] = None
_pbar_lock = Lock()
_counts: Dict[str, int] = {"found": 0, "empty": 0, "failed": 0, "stopped": 0}

_current_lock = Lock()
_current_wallets: set[str] = set()

# Минималистичный pip-подобный формат:
#   🟦 Dune Base  12.2% ━━━━━╸            6/49 │ ⏱ 00:42<05:01 │ 0xD72d9A…A8C7 │ ✓0 ○5 ✗1
_BAR_FORMAT = (
    "{desc}  "
    "{percentage:5.1f}% "
    "{bar:32} "
    "{n_fmt}/{total_fmt} │ "
    "⏱ {elapsed}<{remaining} │ "
    "{current} │ "
    "{stats}"
)
_BAR_ASCII = " ╸━"  # empty · half · full — «пип-стиль»

_CURRENT_WIDTH = 15  # ширина поля current (0xabcdef…wxyz = 12) + запас


def _format_current() -> str:
    with _current_lock:
        if not _current_wallets:
            return "—".ljust(_CURRENT_WIDTH)
        addrs = list(_current_wallets)
        a = addrs[0]
        short = f"{a[:6]}…{a[-4:]}" if len(a) >= 10 else a
        if len(addrs) > 1:
            short = f"{short} +{len(addrs) - 1}"
        return short.ljust(_CURRENT_WIDTH)


def _format_stats() -> str:
    return f"✓{_counts['found']} ○{_counts['empty']} ✗{_counts['failed']}"


class _DuneTqdm(tqdm):
    """tqdm с дополнительными полями ``{current}`` и ``{stats}`` в bar_format."""

    @property
    def format_dict(self) -> Dict[str, Any]:
        d = super().format_dict
        d["current"] = _format_current()
        d["stats"] = _format_stats()
        return d


def _pbar_tick(kind: str) -> None:
    """Увеличивает счётчик ``kind`` и двигает прогресс-бар (потокобезопасно)."""
    with _pbar_lock:
        _counts[kind] = _counts.get(kind, 0) + 1
        if _pbar is not None:
            _pbar.update(1)


def _wallet_begin(address: str) -> None:
    with _current_lock:
        _current_wallets.add(address)
    if _pbar is not None:
        try:
            _pbar.refresh()
        except Exception:  # noqa: BLE001
            pass


def _wallet_end(address: str) -> None:
    with _current_lock:
        _current_wallets.discard(address)


# ──────────────────────────────────────────────────────────────────────────
# Утилиты
# ──────────────────────────────────────────────────────────────────────────

def _derive_address(private_key: str) -> Optional[str]:
    pk = (private_key or "").strip()
    if not pk:
        return None
    if not pk.startswith("0x"):
        pk = "0x" + pk
    try:
        return Account.from_key(pk).address
    except Exception as exc:  # noqa: BLE001
        logger.error(f"Не удалось получить адрес из private_key: {exc}")
        return None


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


# ──────────────────────────────────────────────────────────────────────────
# Обработка одного кошелька
# ──────────────────────────────────────────────────────────────────────────

def _process_wallet(
    scraper: DuneBaseScraper,
    index: int,
    total: int,
    address: str,
    proxy: Optional[str],
    account_name: Optional[str],
) -> Dict[str, Any]:
    max_retries = max(RETRY_COUNT, 1)
    last_error: Optional[str] = None

    for attempt in range(1, max_retries + 1):
        if _stop_event.is_set():
            update_task_failed(address, "Остановлено пользователем")
            _pbar_tick("stopped")
            return {"address": address, "stopped": True}

        try:
            ranking_row, volume_row = scraper.search(address, proxy)

            found_r = bool(ranking_row)
            found_v = bool(volume_row)
            update_task_success(address, found_r, found_v, ranking_row or {}, volume_row or {})

            if found_r or found_v:
                rank = _safe_str((ranking_row or {}).get("rank_tx")) or "—"
                vol = _safe_str(
                    (ranking_row or {}).get("native_volume_eth")
                    or (volume_row or {}).get("native_volume_eth")
                ) or "—"
                log_wallet_task(
                    address, index, total,
                    f"✅ FOUND │ rank={rank} │ vol={vol}",
                    "success", account_name=account_name,
                )
                _pbar_tick("found")
            else:
                # Большинство кошельков не в топе — не засоряем вывод, считаем только в бар.
                _pbar_tick("empty")

            time.sleep(random.uniform(*SLEEP_BETWEEN_ACTIONS))
            return {"address": address, "ok": True}

        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)[:300]
            if attempt < max_retries:
                delay = random.uniform(*SLEEP_BETWEEN_ACTIONS) * 2
                log_task(
                    index, total,
                    f"⚠️ retry {attempt}/{max_retries} │ {last_error[:90]}",
                    "warning", account_name=account_name,
                )
                time.sleep(delay)
            else:
                update_task_failed(address, last_error or "unknown")
                log_wallet_task(
                    address, index, total,
                    f"❌ FAIL │ {last_error[:140]}", "error",
                    account_name=account_name,
                )
                _pbar_tick("failed")
                return {"address": address, "error": last_error}

    update_task_failed(address, last_error or "unknown")
    _pbar_tick("failed")
    return {"address": address, "error": last_error}


# ──────────────────────────────────────────────────────────────────────────
# Сбор данных, многопоточный запуск
# ──────────────────────────────────────────────────────────────────────────

def _worker_loop(
    queue: "Queue[Tuple[int, int, str, Optional[str], Optional[str]]]",
    stagger_sec: float,
) -> None:
    """Рабочая функция одного потока: свой `DuneBaseScraper` на поток."""
    if stagger_sec > 0:
        time.sleep(stagger_sec)
    try:
        with DuneBaseScraper() as scraper:
            while not _stop_event.is_set():
                try:
                    task = queue.get_nowait()
                except Empty:
                    return
                index, ttl, address, proxy, account_name = task
                _wallet_begin(address)
                try:
                    _process_wallet(scraper, index, ttl, address, proxy, account_name)
                finally:
                    _wallet_end(address)
                    queue.task_done()
    except Exception as exc:  # noqa: BLE001
        logger.error(f"Фатальная ошибка рабочего потока: {exc}")


def _collect_wallets() -> List[Dict[str, Optional[str]]]:
    """Возвращает [{wallet_address, proxy, account_name}, ...] из data.csv.

    Каждый кошелёк сохраняет свой proxy из той же строки CSV.
    """
    rows = load_data()
    if not rows:
        return []

    out: List[Dict[str, Optional[str]]] = []
    seen: set = set()
    for row in rows:
        proxy = (row.get("proxy") or "").strip() or None
        name = (row.get("name") or "").strip() or None

        addr: Optional[str] = None
        pk = (row.get("private_key") or "").strip()
        if pk:
            addr = _derive_address(pk)
        if not addr:
            wa = (row.get("wallet_address") or "").strip()
            # Нормализуем: авто-добавляем 0x и приводим префикс к нижнему регистру.
            if wa:
                if wa[:2].lower() == "0x":
                    wa = "0x" + wa[2:]
                else:
                    wa = "0x" + wa
                body = wa[2:]
                if len(body) == 40 and all(c in "0123456789abcdefABCDEF" for c in body):
                    addr = wa
        if not addr:
            continue
        if addr.lower() in seen:
            continue
        seen.add(addr.lower())
        out.append({"wallet_address": addr, "proxy": proxy, "account_name": name})
    return out


def run_base_checker() -> List[Dict[str, Any]]:
    """Главная точка входа чекера. Возвращает все строки результатов из БД."""
    _stop_event.clear()
    init_database()

    wallets = _collect_wallets()
    if not wallets:
        logger.error("Нет кошельков: проверьте data.csv (поле private_key или wallet_address)")
        return []

    inserted = create_check_tasks(wallets)
    if inserted:
        logger.info(f"В БД добавлено новых задач: {inserted}")

    pending_addrs = set(get_pending_addresses())
    todo = [w for w in wallets if w["wallet_address"] in pending_addrs]

    if not todo:
        if pending_addrs:
            logger.warning(
                "Для текущего data.csv нет ожидающих задач. "
                "В БД есть pending/failed задачи от других наборов данных."
            )
        else:
            logger.success("Все кошельки из текущего data.csv уже проверены!")
        _print_summary()
        return get_all_results()

    total = len(todo)
    threads_count = max(1, min(NUM_THREADS, total))

    logger.info(
        f"Dune Base │ к проверке: {total} │ потоков: {threads_count} │ "
        f"попыток: {RETRY_COUNT}"
    )

    q: "Queue[Tuple[int, int, str, Optional[str], Optional[str]]]" = Queue()
    for idx, w in enumerate(todo, 1):
        q.put((idx, total, w["wallet_address"], w["proxy"], w["account_name"]))

    global _pbar, _counts
    _counts = {"found": 0, "empty": 0, "failed": 0, "stopped": 0}

    # tqdm_safe_logging: loguru-логи идут через tqdm.write, чтобы не рвать бар.
    with tqdm_safe_logging(), _DuneTqdm(
        total=total,
        desc=f"{Fore.CYAN}🟦 Dune Base{Style.RESET_ALL}",
        bar_format=_BAR_FORMAT,
        ascii=_BAR_ASCII,
        colour="cyan",
        dynamic_ncols=True,
        leave=True,
        mininterval=0.2,
    ) as pbar:
        _pbar = pbar

        threads: List[Thread] = []
        for i in range(threads_count):
            stagger = random.uniform(*DELAY_BETWEEN_ACCOUNTS) * i
            t = Thread(
                target=_worker_loop,
                args=(q, stagger),
                name=f"dune-base-w{i + 1}",
                daemon=True,
            )
            t.start()
            threads.append(t)

        try:
            while any(t.is_alive() for t in threads):
                for t in threads:
                    t.join(timeout=0.5)
        except KeyboardInterrupt:
            _stop_event.set()
            logger.warning("Прервано (Ctrl+C) — дожидаюсь потоков…")
            for t in threads:
                t.join()

        _pbar = None

    _print_summary()
    path = export_results_xlsx()
    if path:
        logger.success(f"Результаты сохранены: {path}")

    return get_all_results()


# ──────────────────────────────────────────────────────────────────────────
# Итоговая статистика и экспорт
# ──────────────────────────────────────────────────────────────────────────

def _print_summary() -> None:
    s = get_task_statistics()
    if s["total"] == 0:
        return
    sep = f"{Fore.CYAN}{'═' * 60}{Style.RESET_ALL}"
    print()
    print(sep)
    print(f"{Fore.CYAN}  DUNE BASE CHECKER — РЕЗУЛЬТАТЫ{Style.RESET_ALL}")
    print(sep)
    print(f"  Всего кошельков:    {s['total']}")
    print(f"  Проверено:          {Fore.GREEN}{s['completed']}{Style.RESET_ALL}")
    print(f"  Ожидают:            {Fore.YELLOW}{s['pending']}{Style.RESET_ALL}")
    print(f"  Ошибки:             {Fore.RED}{s['failed']}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'─' * 60}{Style.RESET_ALL}")
    print(f"  ✅ Найдено в Ranking: {Fore.GREEN}{s['found_ranking']}{Style.RESET_ALL}")
    print(f"  ✅ Найдено в Volume:  {Fore.GREEN}{s['found_volume']}{Style.RESET_ALL}")
    print(f"  ✅ Найдено хотя бы где-то: {Fore.GREEN}{s['found_any']}{Style.RESET_ALL}")
    print(sep)
    print()


def print_run_statistics() -> None:
    _print_summary()


def export_results_xlsx(output_path: Optional[str] = None) -> Optional[str]:
    """Экспортирует все результаты из БД в xlsx (`result/dune/`)."""
    rows = get_all_results()
    if not rows:
        logger.warning("Нет данных для экспорта (БД пуста)")
        return None

    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(RESULT_DIR / f"dune_base_{ts}.xlsx")

    ranking_keys: List[str] = []
    volume_keys: List[str] = []

    def _extend(keys: List[str], payload: Any) -> None:
        if not payload:
            return
        if isinstance(payload, str):
            try:
                payload = json.loads(payload)
            except Exception:  # noqa: BLE001
                return
        if isinstance(payload, dict):
            for k in payload:
                if k not in keys:
                    keys.append(k)

    for r in rows:
        _extend(ranking_keys, r.get("ranking_data"))
        _extend(volume_keys, r.get("volume_data"))

    wb = Workbook()
    ws = wb.active
    ws.title = "dune_base"

    base_headers = [
        "№",
        "name",
        "wallet_address",
        "status",
        "found_in_ranking",
        "found_in_volume",
        "attempts",
        "error_message",
        "created_at",
        "updated_at",
        "completed_at",
    ]
    r_headers = [f"R:{k}" for k in ranking_keys]
    v_headers = [f"V:{k}" for k in volume_keys]
    headers = base_headers + r_headers + v_headers

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    success_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def _extract(payload: Any, keys: List[str]) -> List[str]:
        if not payload:
            return ["" for _ in keys]
        if isinstance(payload, str):
            try:
                payload = json.loads(payload)
            except Exception:  # noqa: BLE001
                payload = {}
        if not isinstance(payload, dict):
            return ["" for _ in keys]
        return [_safe_str(payload.get(k)) for k in keys]

    for idx, r in enumerate(rows, 1):
        row_values = [
            idx,
            r.get("account_name") or "",
            r.get("wallet_address") or "",
            r.get("status") or "",
            "+" if r.get("found_in_ranking") else "",
            "+" if r.get("found_in_volume") else "",
            r.get("attempts") or 0,
            r.get("error_message") or "",
            _safe_str(r.get("created_at")),
            _safe_str(r.get("updated_at")),
            _safe_str(r.get("completed_at")),
        ]
        row_values += _extract(r.get("ranking_data"), ranking_keys)
        row_values += _extract(r.get("volume_data"), volume_keys)

        ws.append(row_values)
        row_idx = ws.max_row
        row_fill: Optional[PatternFill] = None
        if r.get("status") == "completed" and (r.get("found_in_ranking") or r.get("found_in_volume")):
            row_fill = success_fill
        elif r.get("status") == "failed":
            row_fill = fail_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if row_fill:
                cell.fill = row_fill

    widths = {
        "№": 6, "name": 18, "wallet_address": 44, "status": 11,
        "found_in_ranking": 16, "found_in_volume": 16,
        "attempts": 9, "error_message": 40,
        "created_at": 20, "updated_at": 20, "completed_at": 20,
    }
    for i, h in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = widths.get(h, 18)
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path
