"""Экспорт результатов запуска в один Excel-файл с несколькими листами."""

from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .database import ProxyCheckerDB

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RESULT_DIR = PROJECT_ROOT / "result" / "proxy"


# Цвета по статусам
_STATUS_FILL = {
    "WORKING":   PatternFill("solid", fgColor="C6EFCE"),
    "PARTIAL":   PatternFill("solid", fgColor="FFEB9C"),
    "BROKEN":    PatternFill("solid", fgColor="FFC7CE"),
    "OK":        PatternFill("solid", fgColor="C6EFCE"),
    "BLOCKED":   PatternFill("solid", fgColor="FFD699"),
    "TIMEOUT":   PatternFill("solid", fgColor="F4CCCC"),
    "TLS_ERROR": PatternFill("solid", fgColor="F4CCCC"),
    "PROXY_DOWN":PatternFill("solid", fgColor="F4CCCC"),
    "ERROR":     PatternFill("solid", fgColor="FFC7CE"),
    "HTTP_ERROR":PatternFill("solid", fgColor="FCE5CD"),
    "RPC_ERROR": PatternFill("solid", fgColor="FCE5CD"),
}

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        try:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        except ValueError:
            length = 10
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), max_width)


def _write_header(ws, headers: List[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def export_run_to_xlsx(db: ProxyCheckerDB, run_id: int, *,
                       out_dir: Path = None) -> Path:
    run = db.fetch_run(run_id)
    tasks = db.fetch_tasks(run_id)
    services = db.fetch_service_results(run_id)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = out_dir or (RESULT_DIR / f"run_{run_id}_{timestamp}")
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx = out_dir / "proxy_report.xlsx"

    wb = Workbook()

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ("Run ID", run.get("id")),
        ("Started",  run.get("started_at")),
        ("Finished", run.get("finished_at")),
        ("Detail level", run.get("level")),
        ("Threads", run.get("threads")),
        ("Total proxies", run.get("total")),
        ("Working",  run.get("working")),
        ("Partial",  run.get("partial")),
        ("Broken",   run.get("broken")),
    ]
    for i, (k, v) in enumerate(summary_rows, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # Топ стран
    countries = Counter([t.get("country") or "Unknown" for t in tasks])
    ws.cell(row=len(summary_rows) + 2, column=1, value="Top countries").font = Font(bold=True)
    for i, (country, cnt) in enumerate(countries.most_common(15), 1):
        ws.cell(row=len(summary_rows) + 2 + i, column=1, value=country)
        ws.cell(row=len(summary_rows) + 2 + i, column=2, value=cnt)

    # Топ заблокированных сервисов
    blocked_counter: Counter = Counter()
    error_counter: Counter = Counter()
    failed_stage_counter: Counter = Counter()
    for r in services:
        if r["status"] == "BLOCKED":
            blocked_counter[r["name"]] += 1
        elif r["status"] not in {"OK"}:
            error_counter[f"{r['name']} ({r['status']})"] += 1
        if r.get("failed_stage"):
            failed_stage_counter[r["failed_stage"]] += 1

    base_row = len(summary_rows) + 2 + len(countries) + 3
    ws.cell(row=base_row, column=1, value="Top BLOCKED services").font = Font(bold=True)
    for i, (name, cnt) in enumerate(blocked_counter.most_common(20), 1):
        ws.cell(row=base_row + i, column=1, value=name)
        ws.cell(row=base_row + i, column=2, value=cnt)

    base_row2 = base_row + len(blocked_counter) + 3
    ws.cell(row=base_row2, column=1, value="Top failed stages").font = Font(bold=True)
    for i, (stage, cnt) in enumerate(failed_stage_counter.most_common(), 1):
        ws.cell(row=base_row2 + i, column=1, value=stage)
        ws.cell(row=base_row2 + i, column=2, value=cnt)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    # ---- Sheet 2: Overview ----
    ws = wb.create_sheet("Overview")
    headers = [
        "#", "Proxy", "Status", "Score %", "OK", "Blocked", "Errors",
        "Avg latency ms", "Jitter ms", "Download kbps",
        "IP", "Country", "City", "ASN/Org",
        "Failed stage", "Error", "Started", "Finished",
    ]
    _write_header(ws, headers)
    for row_i, t in enumerate(tasks, start=2):
        details = _safe_json(t.get("details_json"))
        ok_count = blocked_count = err_count = 0
        for r in services:
            if r["task_id"] == t["id"]:
                if r["status"] == "OK":
                    ok_count += 1
                elif r["status"] == "BLOCKED":
                    blocked_count += 1
                else:
                    err_count += 1
        values = [
            t.get("idx"), t.get("proxy"), t.get("overall"),
            t.get("score"), ok_count, blocked_count, err_count,
            t.get("avg_latency"), t.get("jitter_ms"), t.get("download_kbps"),
            t.get("ip"), t.get("country"), t.get("city"), t.get("asn"),
            t.get("failed_stage"), t.get("error"),
            t.get("started_at"), t.get("finished_at"),
        ]
        for col_i, val in enumerate(values, 1):
            ws.cell(row=row_i, column=col_i, value=val)
        fill = _STATUS_FILL.get(t.get("overall") or "")
        if fill:
            ws.cell(row=row_i, column=3).fill = fill
    _autosize(ws)

    # ---- Sheet 3: Service Details ----
    ws = wb.create_sheet("Services")
    headers = [
        "#", "Proxy", "Category", "Service", "URL", "Status", "HTTP",
        "Latency ms", "DNS", "TCP→Proxy", "Proxy CONNECT", "TLS Target",
        "TTFB", "Download ms", "Bytes", "Failed stage", "Error",
    ]
    _write_header(ws, headers)
    proxy_idx = {t["id"]: t.get("idx") for t in tasks}
    for row_i, r in enumerate(services, start=2):
        values = [
            proxy_idx.get(r["task_id"]), r["proxy"], r["category"], r["name"], r["url"],
            r["status"], r["status_code"], r["latency_ms"],
            r["dns_ms"], r["tcp_proxy_ms"], r["proxy_connect_ms"], r["tls_target_ms"],
            r["ttfb_ms"], r["download_ms"], r["bytes_received"],
            r["failed_stage"], r["error"],
        ]
        for col_i, val in enumerate(values, 1):
            ws.cell(row=row_i, column=col_i, value=val)
        fill = _STATUS_FILL.get(r["status"])
        if fill:
            ws.cell(row=row_i, column=6).fill = fill
    _autosize(ws)

    # ---- Sheet 4: Blocked services per proxy ----
    ws = wb.create_sheet("Blocked")
    _write_header(ws, ["#", "Proxy", "Country", "Blocked services", "Count"])
    blocked_by_task: Dict[int, List[str]] = defaultdict(list)
    for r in services:
        if r["status"] == "BLOCKED":
            blocked_by_task[r["task_id"]].append(r["name"])
    row = 2
    for t in tasks:
        names = blocked_by_task.get(t["id"], [])
        if not names:
            continue
        ws.cell(row=row, column=1, value=t.get("idx"))
        ws.cell(row=row, column=2, value=t.get("proxy"))
        ws.cell(row=row, column=3, value=t.get("country"))
        ws.cell(row=row, column=4, value=", ".join(names))
        ws.cell(row=row, column=5, value=len(names))
        row += 1
    _autosize(ws)

    # ---- Sheet 5: Working proxies (clean list) ----
    ws = wb.create_sheet("Working")
    _write_header(ws, ["Proxy", "Country", "Score %", "Avg latency ms"])
    row = 2
    for t in tasks:
        if t.get("overall") == "WORKING":
            ws.cell(row=row, column=1, value=t.get("proxy"))
            ws.cell(row=row, column=2, value=t.get("country"))
            ws.cell(row=row, column=3, value=t.get("score"))
            ws.cell(row=row, column=4, value=t.get("avg_latency"))
            row += 1
    _autosize(ws)

    # ---- Sheet 6: Errors only ----
    ws = wb.create_sheet("Errors")
    _write_header(ws, ["#", "Proxy", "Service", "Status", "Failed stage", "Error", "Latency ms"])
    row = 2
    for r in services:
        if r["status"] in {"OK", "BLOCKED"}:
            continue
        ws.cell(row=row, column=1, value=proxy_idx.get(r["task_id"]))
        ws.cell(row=row, column=2, value=r["proxy"])
        ws.cell(row=row, column=3, value=r["name"])
        ws.cell(row=row, column=4, value=r["status"])
        ws.cell(row=row, column=5, value=r["failed_stage"])
        ws.cell(row=row, column=6, value=r["error"])
        ws.cell(row=row, column=7, value=r["latency_ms"])
        fill = _STATUS_FILL.get(r["status"])
        if fill:
            ws.cell(row=row, column=4).fill = fill
        row += 1
    _autosize(ws)

    wb.save(xlsx)
    return xlsx


def _safe_json(s) -> Dict[str, Any]:
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        return {}
