"""Экспорт результатов Abstract Portal в XLSX."""
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from modules.abs import abs_logger as logger
from modules.abs import database as db

RESULT_DIR = Path(__file__).parent.parent.parent / "result" / "abs_portal"

# ---------------------------------------------------------------
# Стили
# ---------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Fills
_SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зелёный
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # Красный
_WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")      # Жёлтый
_INFO_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")      # Голубой
_GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")       # Золотой
_TIER_FILLS = {
    "bronze": PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid"),
    "silver": PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"),
    "gold": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    "platinum": PatternFill(start_color="E5E4E2", end_color="E5E4E2", fill_type="solid"),
    "diamond": PatternFill(start_color="B9F2FF", end_color="B9F2FF", fill_type="solid"),
}
_TIER_FONT = Font(bold=True, color="000000", size=11)
_XP_FONT = Font(bold=True, color="2E7D32", size=11)
_NUM_FONT = Font(bold=False, color="333333", size=10)


def _auto_width(ws, col_count, row_count, max_w=50):
    """Установить автоширину колонок."""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in range(1, min(row_count + 1, 200)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)), max_w))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 3


def _write_headers(ws, headers, row=1):
    """Записать заголовки с форматированием."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _get_tier_fill(tier_name: str) -> PatternFill | None:
    """Получить цвет для тира."""
    name_lower = (tier_name or "").lower()
    for key, fill in _TIER_FILLS.items():
        if key in name_lower:
            return fill
    return None


# ---------------------------------------------------------------
# Экспорт
# ---------------------------------------------------------------

def export_stats(filename: str = None) -> str | None:
    """Экспортировать полную статистику Abstract Portal в XLSX.

    Листы:
      1. Profile   — профиль каждого кошелька (ранг, XP, streak, socials)
      2. Badges    — бейджи (заклеймленные / доступные для клейма)
      3. XP Recap  — XP по неделям/эпохам
      4. Summary   — сводная статистика
    """
    wallets = db.get_all_wallets()
    if not wallets:
        logger.log("Нет данных для экспорта", "warning")
        return None

    os.makedirs(RESULT_DIR, exist_ok=True)

    if filename is None:
        filename = f"abs_portal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = str(RESULT_DIR / filename)

    try:
        wb = Workbook()

        # ═══════════════════════════════════════════════
        # Лист 1: Profile
        # ═══════════════════════════════════════════════
        ws = wb.active
        ws.title = "Profile"

        headers = [
            "#", "Account", "EVM Address", "ABS Wallet",
            "Rank", "Total XP", "XP Multiplier",
            "Streak (Days)", "Longest Streak", "Voted Today",
            "Twitter", "Discord",
            "Badges Claimed", "ETH Balance",
            "Last Updated",
        ]
        _write_headers(ws, headers)

        for idx, w in enumerate(wallets, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx).border = _THIN_BORDER

            # Account name
            ws.cell(row=row, column=2,
                    value=w.get("account_name") or f"Wallet_{idx}").border = _THIN_BORDER

            # EVM Address
            ws.cell(row=row, column=3, value=w.get("address", "")).border = _THIN_BORDER

            # ABS Wallet
            ws.cell(row=row, column=4,
                    value=w.get("abs_wallet_address") or "").border = _THIN_BORDER

            # Rank (tier_name) — с цветом
            tier_name = w.get("tier_name") or ""
            tier_cell = ws.cell(row=row, column=5, value=tier_name)
            tier_cell.border = _THIN_BORDER
            tier_cell.font = _TIER_FONT
            tier_fill = _get_tier_fill(tier_name)
            if tier_fill:
                tier_cell.fill = tier_fill

            # Total XP
            xp_cell = ws.cell(row=row, column=6, value=w.get("total_xp", 0))
            xp_cell.border = _THIN_BORDER
            xp_cell.font = _XP_FONT
            xp_cell.alignment = Alignment(horizontal="right")

            # XP Multiplier
            mult = w.get("xp_multiplier", 1.0)
            mult_cell = ws.cell(row=row, column=7, value=f"{mult}x")
            mult_cell.border = _THIN_BORDER
            if mult and mult > 1:
                mult_cell.fill = _GOLD_FILL

            # Streak
            streak_cell = ws.cell(row=row, column=8,
                                  value=w.get("current_streak_days", 0))
            streak_cell.border = _THIN_BORDER
            if w.get("current_streak_days", 0) >= 7:
                streak_cell.fill = _SUCCESS_FILL

            ws.cell(row=row, column=9,
                    value=w.get("longest_streak_days", 0)).border = _THIN_BORDER

            # Voted Today
            voted = w.get("voted_today", 0)
            voted_cell = ws.cell(row=row, column=10,
                                 value="Yes" if voted else "No")
            voted_cell.border = _THIN_BORDER
            voted_cell.fill = _SUCCESS_FILL if voted else _FAIL_FILL

            # Twitter
            tw = w.get("twitter_connected", 0)
            tw_cell = ws.cell(row=row, column=11,
                              value="Connected" if tw else "No")
            tw_cell.border = _THIN_BORDER
            tw_cell.fill = _SUCCESS_FILL if tw else _WARN_FILL

            # Discord
            dc = w.get("discord_connected", 0)
            dc_cell = ws.cell(row=row, column=12,
                              value="Connected" if dc else "No")
            dc_cell.border = _THIN_BORDER
            dc_cell.fill = _SUCCESS_FILL if dc else _WARN_FILL

            # Badges claimed count
            badges = db.get_badges(w["address"])
            claimed = sum(1 for b in badges if b.get("claimed"))
            badge_cell = ws.cell(row=row, column=13,
                                 value=f"{claimed}/{len(badges)}")
            badge_cell.border = _THIN_BORDER
            if badges and claimed == len(badges):
                badge_cell.fill = _SUCCESS_FILL
            elif claimed > 0:
                badge_cell.fill = _WARN_FILL
            else:
                badge_cell.fill = _FAIL_FILL

            # ETH Balance
            ws.cell(row=row, column=14,
                    value=w.get("eth_balance") or "N/A").border = _THIN_BORDER

            # Last Updated
            ws.cell(row=row, column=15,
                    value=(w.get("stats_updated_at") or "N/A")[:19]).border = _THIN_BORDER

        _auto_width(ws, len(headers), len(wallets) + 1)

        # ═══════════════════════════════════════════════
        # Лист 2: Badges
        # ═══════════════════════════════════════════════
        ws2 = wb.create_sheet("Badges")

        # Собираем все бейджи по всем кошелькам
        badge_headers = [
            "#", "Account", "Address", "Badge", "Description",
            "Requirement", "Type", "Status",
        ]
        _write_headers(ws2, badge_headers)

        badge_row = 2
        for idx, w in enumerate(wallets, 1):
            badges = db.get_badges(w["address"])
            acct_name = w.get("account_name") or f"Wallet_{idx}"
            for b in badges:
                ws2.cell(row=badge_row, column=1, value=badge_row - 1).border = _THIN_BORDER
                ws2.cell(row=badge_row, column=2, value=acct_name).border = _THIN_BORDER
                ws2.cell(row=badge_row, column=3, value=w["address"]).border = _THIN_BORDER
                ws2.cell(row=badge_row, column=4,
                         value=b.get("badge_name") or "").border = _THIN_BORDER
                ws2.cell(row=badge_row, column=5,
                         value=b.get("badge_description") or "").border = _THIN_BORDER
                ws2.cell(row=badge_row, column=6,
                         value=b.get("badge_requirement") or "").border = _THIN_BORDER
                ws2.cell(row=badge_row, column=7,
                         value=b.get("badge_type") or "regular").border = _THIN_BORDER

                # Status with color
                if b.get("claimed"):
                    status = "Claimed"
                    fill = _SUCCESS_FILL
                elif b.get("claim_available"):
                    status = "Available"
                    fill = _GOLD_FILL
                else:
                    status = "Locked"
                    fill = _FAIL_FILL

                status_cell = ws2.cell(row=badge_row, column=8, value=status)
                status_cell.border = _THIN_BORDER
                status_cell.fill = fill
                status_cell.font = Font(bold=True)

                badge_row += 1

        _auto_width(ws2, len(badge_headers), badge_row)

        # ═══════════════════════════════════════════════
        # Лист 2.5: Badge Summary (сводка по бейджам)
        # ═══════════════════════════════════════════════
        ws_bs = wb.create_sheet("Badge Summary")

        bs_headers = [
            "#", "Account", "Address",
            "Claimed (qty)", "Claimed Badges",
            "Claimable (qty)", "Claimable Badges",
            "Not Done (qty)", "Not Done Badges",
        ]
        _write_headers(ws_bs, bs_headers)

        # Цвета для ячеек с названиями бейджей
        _GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        _GREEN_FONT = Font(bold=True, color="006100", size=10)
        _YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        _YELLOW_FONT = Font(bold=True, color="9C6500", size=10)
        _RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        _RED_FONT = Font(bold=True, color="9C0006", size=10)

        for idx, w in enumerate(wallets, 1):
            bs_row = idx + 1
            badges = db.get_badges(w["address"])
            acct_name = w.get("account_name") or f"Wallet_{idx}"

            claimed_names = [b["badge_name"] for b in badges if b.get("claimed")]
            claimable_names = [b["badge_name"] for b in badges if not b.get("claimed") and b.get("claim_available")]
            not_done_names = [b["badge_name"] for b in badges if not b.get("claimed") and not b.get("claim_available")]

            ws_bs.cell(row=bs_row, column=1, value=idx).border = _THIN_BORDER
            ws_bs.cell(row=bs_row, column=2, value=acct_name).border = _THIN_BORDER
            ws_bs.cell(row=bs_row, column=3, value=w["address"]).border = _THIN_BORDER

            # Claimed — зелёный
            c_qty = ws_bs.cell(row=bs_row, column=4, value=len(claimed_names))
            c_qty.border = _THIN_BORDER
            c_qty.fill = _GREEN_FILL
            c_qty.font = _GREEN_FONT
            c_names = ws_bs.cell(row=bs_row, column=5, value=", ".join(claimed_names) if claimed_names else "—")
            c_names.border = _THIN_BORDER
            c_names.fill = _GREEN_FILL
            c_names.font = _GREEN_FONT

            # Claimable — жёлтый
            a_qty = ws_bs.cell(row=bs_row, column=6, value=len(claimable_names))
            a_qty.border = _THIN_BORDER
            a_qty.fill = _YELLOW_FILL
            a_qty.font = _YELLOW_FONT
            a_names = ws_bs.cell(row=bs_row, column=7, value=", ".join(claimable_names) if claimable_names else "—")
            a_names.border = _THIN_BORDER
            a_names.fill = _YELLOW_FILL
            a_names.font = _YELLOW_FONT

            # Not done — красный
            n_qty = ws_bs.cell(row=bs_row, column=8, value=len(not_done_names))
            n_qty.border = _THIN_BORDER
            n_qty.fill = _RED_FILL
            n_qty.font = _RED_FONT
            n_names = ws_bs.cell(row=bs_row, column=9, value=", ".join(not_done_names) if not_done_names else "—")
            n_names.border = _THIN_BORDER
            n_names.fill = _RED_FILL
            n_names.font = _RED_FONT

        _auto_width(ws_bs, len(bs_headers), len(wallets) + 1)

        # ═══════════════════════════════════════════════
        # Лист 2.6: Badge Matrix (универсальная матрица бейджей)
        # Колонки = все уникальные бейджи из БД, строки = кошельки
        # Зелёный = заклеймлен, Жёлтый = готов к клейму, Красный = не готов
        # ═══════════════════════════════════════════════
        ws_bm = wb.create_sheet("Badge Matrix")

        all_unique_badges = db.get_all_unique_badges()

        bm_headers = ["#", "Account", "Address"]
        for ub in all_unique_badges:
            bm_headers.append(ub["badge_name"] or f"Badge #{ub['badge_id']}")
        _write_headers(ws_bm, bm_headers)

        _CLAIMED_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        _CLAIMED_FONT = Font(bold=True, color="FFFFFF", size=10)
        _CLAIMABLE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        _CLAIMABLE_FONT = Font(bold=True, color="000000", size=10)
        _NOT_READY_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        _NOT_READY_FONT = Font(bold=True, color="FFFFFF", size=10)

        for idx, w in enumerate(wallets, 1):
            bm_row = idx + 1
            ws_bm.cell(row=bm_row, column=1, value=idx).border = _THIN_BORDER
            ws_bm.cell(row=bm_row, column=2,
                        value=w.get("account_name") or f"Wallet_{idx}").border = _THIN_BORDER
            ws_bm.cell(row=bm_row, column=3, value=w["address"]).border = _THIN_BORDER

            # Получить бейджи этого кошелька, индексировать по badge_id
            wallet_badges = db.get_badges(w["address"])
            badges_by_id = {b["badge_id"]: b for b in wallet_badges}

            for col_offset, ub in enumerate(all_unique_badges):
                col = col_offset + 4
                b = badges_by_id.get(ub["badge_id"])

                if b and b.get("claimed"):
                    status_text = "Claimed"
                    fill = _CLAIMED_FILL
                    font = _CLAIMED_FONT
                elif b and b.get("claim_available"):
                    status_text = "Claimable"
                    fill = _CLAIMABLE_FILL
                    font = _CLAIMABLE_FONT
                else:
                    status_text = "Not Ready"
                    fill = _NOT_READY_FILL
                    font = _NOT_READY_FONT

                cell = ws_bm.cell(row=bm_row, column=col, value=status_text)
                cell.border = _THIN_BORDER
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")

        _auto_width(ws_bm, len(bm_headers), len(wallets) + 1)

        # ═══════════════════════════════════════════════
        # Лист 3: XP Recap
        # ═══════════════════════════════════════════════
        ws3 = wb.create_sheet("XP Recap")

        xp_headers = [
            "#", "Account", "Address", "Epoch (Week)",
            "Description", "XP Earned", "Referral XP", "Season",
        ]
        _write_headers(ws3, xp_headers)

        xp_row = 2
        for idx, w in enumerate(wallets, 1):
            xp_items = db.get_xp_history(w["address"])
            acct_name = w.get("account_name") or f"Wallet_{idx}"
            for item in xp_items:
                ws3.cell(row=xp_row, column=1, value=xp_row - 1).border = _THIN_BORDER
                ws3.cell(row=xp_row, column=2, value=acct_name).border = _THIN_BORDER
                ws3.cell(row=xp_row, column=3, value=w["address"]).border = _THIN_BORDER
                ws3.cell(row=xp_row, column=4,
                         value=item.get("epoch", 0)).border = _THIN_BORDER
                ws3.cell(row=xp_row, column=5,
                         value=item.get("description") or "").border = _THIN_BORDER

                # XP Earned — зелёный если > 0
                xp_val = item.get("points", 0)
                xp_cell = ws3.cell(row=xp_row, column=6, value=xp_val)
                xp_cell.border = _THIN_BORDER
                xp_cell.font = _XP_FONT if xp_val > 0 else _NUM_FONT
                if xp_val > 0:
                    xp_cell.fill = _SUCCESS_FILL

                # Referral XP
                ref_xp = item.get("referral_xp", 0)
                ref_cell = ws3.cell(row=xp_row, column=7, value=ref_xp)
                ref_cell.border = _THIN_BORDER
                if ref_xp > 0:
                    ref_cell.fill = _INFO_FILL

                ws3.cell(row=xp_row, column=8,
                         value=item.get("season", 0)).border = _THIN_BORDER

                xp_row += 1

        _auto_width(ws3, len(xp_headers), xp_row)

        # ═══════════════════════════════════════════════
        # Лист 4: Summary
        # ═══════════════════════════════════════════════
        ws4 = wb.create_sheet("Summary")

        stats = db.get_db_stats()

        # Общая сводка
        _write_headers(ws4, ["Parameter", "Value"])

        total_xp_all = sum(w.get("total_xp", 0) for w in wallets)
        avg_xp = total_xp_all // len(wallets) if wallets else 0

        # Подсчёт по тирам
        tier_counts: dict[str, int] = {}
        for w in wallets:
            t = w.get("tier_name") or "Unknown"
            tier_counts[t] = tier_counts.get(t, 0) + 1

        summary_data = [
            ("Total Wallets", stats.get("total_wallets", 0)),
            ("Logged In", stats.get("logged_in", 0)),
            ("Stats Collected", stats.get("with_stats", 0)),
            ("", ""),
            ("Total XP (all wallets)", total_xp_all),
            ("Average XP per wallet", avg_xp),
            ("", ""),
            ("Unique Badges", stats.get("unique_badges", 0)),
            ("Total Badges Claimed", stats.get("claimed_badges", 0)),
            ("", ""),
        ]

        # Tier distribution
        summary_data.append(("--- Tier Distribution ---", ""))
        for tier, count in sorted(tier_counts.items()):
            summary_data.append((tier, count))

        summary_data.append(("", ""))
        summary_data.append(("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

        for idx, (param, val) in enumerate(summary_data, 1):
            param_cell = ws4.cell(row=idx + 1, column=1, value=param)
            param_cell.border = _THIN_BORDER
            if param.startswith("---"):
                param_cell.font = Font(bold=True, color="2E7D32")
            elif param == "":
                pass
            else:
                param_cell.font = Font(bold=False)

            val_cell = ws4.cell(row=idx + 1, column=2, value=val)
            val_cell.border = _THIN_BORDER
            val_cell.font = Font(bold=True)

        _auto_width(ws4, 2, len(summary_data) + 1)

        wb.save(filepath)
        logger.log(f"Экспорт сохранён: {filepath}", "success")
        return filepath

    except Exception as e:
        logger.log(f"Ошибка экспорта XLSX: {e}", "error")
        return None
