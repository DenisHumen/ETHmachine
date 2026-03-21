"""Pharos Discord Connect — многопоточный обработчик задач."""
import random
import time
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Event

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from eth_account import Account

from config.modules.cfg_base import (
    NUM_THREADS, RETRY_COUNT,
    SLEEP_BETWEEN_ACTIONS, DELAY_BETWEEN_ACCOUNTS,
)
from config.modules.cfg_pharos_discord import RESULT_DIR
from modules.simple_logger import log_wallet_task, log_simple
from modules.proxy_manager import parse_proxy, mask_proxy
from modules.data_manager import load_data
from modules.pharos.discord_connect import database as db
from modules.pharos.discord_connect.client import PharosDiscordClient

RESULT_PATH = Path(RESULT_DIR)
_stop_event = Event()

# Runtime-параметры (переопределяются при запуске из меню)
_runtime_retry_count: int = RETRY_COUNT
_runtime_sleep_between_actions: list = SLEEP_BETWEEN_ACTIONS
_runtime_delay_between_accounts: list = DELAY_BETWEEN_ACCOUNTS


# ─────────────────── DATA LOADING ───────────────────

def load_wallets_from_csv() -> list[dict]:
    """Загрузить кошельки, прокси и Discord токены из data.csv."""
    rows = load_data()
    if not rows:
        log_simple("Нет данных в data.csv!", "error")
        return []

    wallets = []
    for i, row in enumerate(rows):
        pk = row.get("private_key", "").strip()
        if not pk:
            continue
        if not pk.startswith("0x"):
            pk = "0x" + pk
        try:
            account = Account.from_key(pk)
            proxy = row.get("proxy", "").strip() or None
            discord_token = row.get("discord_token", "").strip() or None
            account_name = row.get("name", "").strip() or None
            wallets.append({
                "private_key": pk,
                "address": account.address,
                "account_name": account_name,
                "proxy": proxy,
                "discord_token": discord_token,
            })
        except Exception as e:
            log_simple(f"Ошибка ключа #{i + 1}: {e}", "error")

    return wallets


def create_tasks_from_csv() -> int:
    """Загрузить данные из CSV и создать задачи в БД. Возвращает количество."""
    wallets = load_wallets_from_csv()
    if not wallets:
        return 0

    db.create_tasks(wallets)
    return len(wallets)


# ─────────────────── SINGLE WALLET PROCESSING ───────────────────

def _process_single_wallet(
    index: int,
    total: int,
    task: dict,
) -> dict:
    """Обработать один кошелёк: авторизация + привязка Discord."""
    address = task["address"]
    private_key = task["private_key"]
    proxy = task.get("proxy")
    discord_token = task.get("discord_token")
    account_name = task.get("account_name")

    retry_count = _runtime_retry_count
    sleep_actions = _runtime_sleep_between_actions

    result = {
        "address": address,
        "status": "failed",
        "discord_username": None,
        "discord_id": None,
        "error": None,
    }

    if not discord_token:
        msg = "Discord токен не указан в data.csv"
        log_wallet_task(address, index, total, msg, "warning", account_name=account_name)
        db.update_task_failed(address, msg)
        result["error"] = msg
        result["status"] = "skipped"
        return result

    proxy_display = mask_proxy(proxy) if proxy else "без прокси"
    log_wallet_task(address, index, total, f"Старт | Прокси: {proxy_display}", "info", account_name=account_name)

    client = PharosDiscordClient(
        private_key=private_key,
        proxy=proxy,
        discord_token=discord_token,
        wallet_index=index - 1,
    )

    def _log(msg, level="info"):
        log_wallet_task(address, index, total, msg, level, account_name=account_name)

    client.set_logger(_log)

    for attempt in range(1, retry_count + 1):
        if _stop_event.is_set():
            result["error"] = "Остановлено пользователем"
            return result

        try:
            success = client.connect_discord()
            if success:
                result["status"] = "completed"
                task_data = db.get_task_by_address(address)
                if task_data:
                    result["discord_username"] = task_data.get("discord_username")
                    result["discord_id"] = task_data.get("discord_id")
                log_wallet_task(address, index, total, "Discord привязан!", "success", account_name=account_name)
                return result
            else:
                if attempt < retry_count:
                    delay = random.uniform(*sleep_actions)
                    log_wallet_task(
                        address, index, total,
                        f"Попытка {attempt}/{retry_count} неудачна, ждём {delay:.1f}с...",
                        "warning",
                        account_name=account_name,
                    )
                    time.sleep(delay)
        except Exception as e:
            error_msg = str(e)[:100]
            if attempt < retry_count:
                log_wallet_task(
                    address, index, total,
                    f"Ошибка попытки {attempt}/{retry_count}: {error_msg}",
                    "error",
                    account_name=account_name,
                )
                time.sleep(random.uniform(*sleep_actions))
            else:
                db.update_task_failed(address, error_msg)
                result["error"] = error_msg

    log_wallet_task(address, index, total, "Все попытки исчерпаны", "error", account_name=account_name)
    return result


def _thread_worker(args: tuple) -> dict:
    """Обёртка для ThreadPoolExecutor."""
    index, total, task = args

    # Задержка между стартами аккаунтов
    if index > 1:
        delay = random.uniform(*_runtime_delay_between_accounts)
        time.sleep(delay)

    return _process_single_wallet(index, total, task)


# ─────────────────── MAIN RUN ───────────────────

def run_discord_connect(
    max_workers: int | None = None,
    retry_count: int | None = None,
    sleep_between_actions: list | None = None,
    delay_between_accounts: list | None = None,
):
    """Запуск привязки Discord для всех pending задач.

    Args:
        max_workers: количество потоков (по умолчанию NUM_THREADS из cfg_base)
        retry_count: количество ретраев (по умолчанию RETRY_COUNT из cfg_base)
        sleep_between_actions: задержка между действиями [min, max] сек
        delay_between_accounts: задержка между стартом аккаунтов [min, max] сек
    """
    global _runtime_retry_count, _runtime_sleep_between_actions, _runtime_delay_between_accounts

    _runtime_retry_count = retry_count if retry_count is not None else RETRY_COUNT
    _runtime_sleep_between_actions = sleep_between_actions if sleep_between_actions is not None else SLEEP_BETWEEN_ACTIONS
    _runtime_delay_between_accounts = delay_between_accounts if delay_between_accounts is not None else DELAY_BETWEEN_ACCOUNTS

    _stop_event.clear()

    all_tasks = db.get_all_tasks()
    if not all_tasks:
        log_simple("Нет задач в базе данных. Создайте задачи из CSV.", "error")
        return

    pending = [t for t in all_tasks if t["status"] in ("pending", "failed", "auth_ok")]
    if not pending:
        log_simple("Нет pending/failed задач. Все задачи выполнены!", "success")
        return

    total = len(pending)
    threads = max_workers or min(NUM_THREADS, total)

    log_simple(
        f"Запуск: {total} задач, {threads} потоков, "
        f"{_runtime_retry_count} ретраев, "
        f"задержка дей: {_runtime_sleep_between_actions[0]}-{_runtime_sleep_between_actions[1]}с, "
        f"задержка акк: {_runtime_delay_between_accounts[0]}-{_runtime_delay_between_accounts[1]}с",
        "info",
    )

    # Подготовка аргументов
    tasks_args = [(i + 1, total, task) for i, task in enumerate(pending)]

    results = []
    try:
        with ThreadPoolExecutor(max_workers=threads) as executor:
            futures = {executor.submit(_thread_worker, args): args for args in tasks_args}
            for future in as_completed(futures):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    log_simple(f"Ошибка потока: {e}", "error")
    except KeyboardInterrupt:
        _stop_event.set()
        log_simple("Остановлено пользователем (Ctrl+C)", "warning")

    # Итоги
    completed = sum(1 for r in results if r["status"] == "completed")
    failed = sum(1 for r in results if r["status"] == "failed")
    skipped = sum(1 for r in results if r["status"] == "skipped")

    log_simple(
        f"Итого: {completed} успешно, {failed} ошибок, {skipped} пропущено из {total}",
        "success" if failed == 0 else "warning",
    )


# ─────────────────── EXCEL EXPORT ───────────────────

def export_results_xlsx(output_path: str | None = None) -> str | None:
    """Экспорт результатов в Excel файл."""
    results = db.get_all_results()
    if not results:
        log_simple("Нет результатов для экспорта", "warning")
        return None

    RESULT_PATH.mkdir(parents=True, exist_ok=True)

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(RESULT_PATH / f"pharos_discord_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pharos Discord Connect"

    # Стили
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Consolas", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    status_colors = {
        "completed": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "failed": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "pending": PatternFill(start_color="D6D8DB", end_color="D6D8DB", fill_type="solid"),
        "auth_ok": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "skipped": PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"),
    }

    # Заголовки
    headers = [
        "#", "Address", "Status", "Discord Username", "Discord ID",
        "Proxy", "Error", "Attempts", "Created", "Completed",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Данные
    for idx, row_data in enumerate(results, 1):
        row_num = idx + 1
        proxy_display = mask_proxy(row_data.get("proxy")) if row_data.get("proxy") else ""
        values = [
            idx,
            row_data.get("address", ""),
            row_data.get("status", ""),
            row_data.get("discord_username", "") or "",
            row_data.get("discord_id", "") or "",
            proxy_display,
            row_data.get("error_message", "") or "",
            row_data.get("attempts", 0),
            row_data.get("created_at", ""),
            row_data.get("completed_at", "") or "",
        ]

        fill = status_colors.get(row_data.get("status", ""), PatternFill())

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    # Ширина колонок
    col_widths = [5, 44, 12, 22, 20, 22, 40, 8, 20, 20]
    for i, width in enumerate(col_widths, 1):
        letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)
        ws.column_dimensions[letter].width = width

    # Итоговая строка
    total_row = len(results) + 2
    completed_count = sum(1 for r in results if r.get("status") == "completed")
    failed_count = sum(1 for r in results if r.get("status") == "failed")
    pending_count = sum(1 for r in results if r.get("status") == "pending")

    summary_font = Font(name="Consolas", bold=True, size=11)
    ws.cell(row=total_row, column=1, value="ИТОГО").font = summary_font
    ws.cell(row=total_row, column=2, value=f"Всего: {len(results)}").font = summary_font
    ws.cell(row=total_row, column=3, value=f"OK: {completed_count}").font = Font(
        name="Consolas", bold=True, color="28A745", size=11
    )
    ws.cell(row=total_row, column=4, value=f"Fail: {failed_count}").font = Font(
        name="Consolas", bold=True, color="DC3545", size=11
    )
    ws.cell(row=total_row, column=5, value=f"Pending: {pending_count}").font = Font(
        name="Consolas", bold=True, color="6C757D", size=11
    )

    wb.save(output_path)
    log_simple(f"Результаты сохранены: {output_path}", "success")
    return output_path
