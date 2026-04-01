"""Экспорт результатов Pharos в XLSX — из БД и из памяти."""
import json
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from modules.pharos import pharos_logger as logger
from modules.pharos import database as db

RESULT_DIR = Path(__file__).parent.parent.parent / "result" / "pharos_auto"

# ═══════════════════════════════════════════════════════════
# Общие стили
# ═══════════════════════════════════════════════════════════

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_LABEL_FONT = Font(bold=True, size=11)


def _auto_width(ws, col_count, row_count, max_w=50):
    """Установить автоширину колонок."""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in range(1, row_count + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)), max_w))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 3


def _write_headers(ws, headers, row=1):
    """Записать заголовки с форматированием."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


# ═══════════════════════════════════════════════════════════
# Экспорт результатов цикла из БД
# ═══════════════════════════════════════════════════════════

def export_cycle_results(results_table: str, filename: str = None) -> str | None:
    """Экспортировать результаты завершённого цикла из таблицы БД в XLSX.

    Args:
        results_table: имя таблицы cycle_results_* в БД
        filename: имя выходного файла (по умолчанию авто)

    Returns:
        путь к файлу или None
    """
    rows = db.get_cycle_results(results_table)
    if not rows:
        logger.log(f"Нет результатов в таблице {results_table}", "warning")
        return None

    os.makedirs(RESULT_DIR, exist_ok=True)

    if filename is None:
        filename = f"{results_table}.xlsx"
    filepath = str(RESULT_DIR / filename)

    try:
        wb = Workbook()

        # ─── Лист 1: Результаты ───
        ws = wb.active
        ws.title = "Results"

        headers = [
            "№", "Account Name", "Address", "Status",
            "Started", "Finished", "Details", "Error",
        ]
        _write_headers(ws, headers)

        success_count = 0
        failed_count = 0
        all_temp_wallets = []

        for i, r in enumerate(rows, 1):
            row_num = i + 1
            status = r.get("status", "unknown")
            is_ok = status == "completed"
            if is_ok:
                success_count += 1
            else:
                failed_count += 1

            # Парсить result_json для деталей
            details = ""
            result_data = {}
            if r.get("result_json"):
                try:
                    result_data = json.loads(r["result_json"])
                    parts = []
                    if result_data.get("sends"):
                        parts.append(f"Sends: {result_data['sends']}")
                    if result_data.get("verifies"):
                        parts.append(f"Verifies: {result_data['verifies']}")
                    if result_data.get("xp_gained"):
                        parts.append(f"XP: +{result_data['xp_gained']}")
                    if result_data.get("faucet_claimed"):
                        parts.append("Faucet: Yes")
                    if result_data.get("checkin_done"):
                        parts.append("Check-in: Yes")
                    if result_data.get("balance_before") is not None:
                        parts.append(f"Balance: {result_data.get('balance_before', 0):.6f} -> {result_data.get('balance_after', 0):.6f}")
                    details = " | ".join(parts)
                except (json.JSONDecodeError, TypeError):
                    details = str(r.get("result_json", ""))[:100]

            # Temp wallets
            for tw in result_data.get("temp_wallets", []):
                all_temp_wallets.append({
                    "main_address": r.get("wallet_address", ""),
                    "account_name": r.get("account_name", ""),
                    **tw,
                })

            values = [
                i,
                r.get("account_name", ""),
                r.get("wallet_address", ""),
                "OK" if is_ok else "FAIL",
                r.get("started_at", ""),
                r.get("finished_at", ""),
                details,
                r.get("error", ""),
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = _THIN_BORDER
                if col == 4:
                    cell.fill = _SUCCESS_FILL if is_ok else _FAIL_FILL
                    cell.alignment = Alignment(horizontal="center")

        _auto_width(ws, len(headers), len(rows) + 1)

        # ─── Лист 2: Temp Wallets (если есть) ───
        if all_temp_wallets:
            ws2 = wb.create_sheet("Temp Wallets")
            temp_headers = ["№", "Main Address", "Account Name", "Temp Address", "Temp Private Key"]
            _write_headers(ws2, temp_headers)

            for i, tw in enumerate(all_temp_wallets, 1):
                values = [i, tw["main_address"], tw["account_name"],
                          tw.get("address", ""), tw.get("private_key", "")]
                for col, val in enumerate(values, 1):
                    cell = ws2.cell(row=i + 1, column=col, value=val)
                    cell.border = _THIN_BORDER
            _auto_width(ws2, len(temp_headers), len(all_temp_wallets) + 1, 70)

        # ─── Лист 3: Сводка ───
        ws3 = wb.create_sheet("Summary")
        summary = [
            ("Таблица", results_table),
            ("Дата экспорта", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Всего аккаунтов", len(rows)),
            ("Успешно", success_count),
            ("Неудачно", failed_count),
        ]
        for row_n, (label, value) in enumerate(summary, 1):
            cell_l = ws3.cell(row=row_n, column=1, value=label)
            cell_l.font = _LABEL_FONT
            cell_l.border = _THIN_BORDER
            cell_v = ws3.cell(row=row_n, column=2, value=value)
            cell_v.border = _THIN_BORDER
        ws3.column_dimensions["A"].width = 25
        ws3.column_dimensions["B"].width = 40

        wb.save(filepath)
        logger.log(f"Результаты сохранены: {filepath} ({len(rows)} аккаунтов)", "success")
        return filepath

    except Exception as e:
        logger.log(f"Ошибка экспорта XLSX: {e}", "error")
        return None


# ═══════════════════════════════════════════════════════════
# Экспорт из памяти (для обратной совместимости)
# ═══════════════════════════════════════════════════════════

def export_send_verify_results(results: list[dict], filename: str = None) -> str | None:
    """Экспортировать результаты send_verify из памяти в XLSX."""
    if not results:
        logger.log("Нет результатов для экспорта", "warning")
        return None

    os.makedirs(RESULT_DIR, exist_ok=True)

    if filename is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pharos_auto_{ts}.xlsx"

    filepath = str(RESULT_DIR / filename)

    try:
        wb = Workbook()

        # Лист 1: Результаты
        ws = wb.active
        ws.title = "Results"

        headers = [
            "№", "Account Name", "Address", "Status",
            "XP Before", "XP After", "XP Gained",
            "Balance Before", "Balance After",
            "Sends", "Verifies",
            "Faucet", "Check-in",
            "TX Hashes", "Error",
        ]
        _write_headers(ws, headers)

        for i, r in enumerate(results, 1):
            row = i + 1
            status = "OK" if r.get("success") else "FAIL"
            tx_hashes = "\n".join(r.get("tx_hashes", []))

            values = [
                i, r.get("account_name", ""), r.get("address", ""), status,
                r.get("xp_before", 0), r.get("xp_after", 0), r.get("xp_gained", 0),
                round(r.get("balance_before", 0), 6), round(r.get("balance_after", 0), 6),
                r.get("sends", 0), r.get("verifies", 0),
                "Yes" if r.get("faucet_claimed") else "No",
                "Yes" if r.get("checkin_done") else "No",
                tx_hashes, r.get("error", ""),
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = _THIN_BORDER
                if col == 4:
                    cell.fill = _SUCCESS_FILL if status == "OK" else _FAIL_FILL
                    cell.alignment = Alignment(horizontal="center")

        _auto_width(ws, len(headers), len(results) + 1)

        # Лист 2: Temp Wallets
        ws2 = wb.create_sheet("Temp Wallets")
        temp_headers = ["№", "Main Address", "Account Name", "Temp Address", "Temp Private Key"]
        _write_headers(ws2, temp_headers)

        temp_row = 2
        for r in results:
            for tw in r.get("temp_wallets", []):
                values = [temp_row - 1, r.get("address", ""), r.get("account_name", ""),
                          tw.get("address", ""), tw.get("private_key", "")]
                for col, val in enumerate(values, 1):
                    cell = ws2.cell(row=temp_row, column=col, value=val)
                    cell.border = _THIN_BORDER
                temp_row += 1
        _auto_width(ws2, len(temp_headers), temp_row - 1, 70)

        # Лист 3: Сводка
        ws3 = wb.create_sheet("Summary")
        total = len(results)
        summary = [
            ("Дата", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Всего аккаунтов", total),
            ("Успешно", sum(1 for r in results if r.get("success"))),
            ("Неудачно", sum(1 for r in results if not r.get("success"))),
            ("Отправок TX", sum(r.get("sends", 0) for r in results)),
            ("Верификаций", sum(r.get("verifies", 0) for r in results)),
            ("XP получено", sum(r.get("xp_gained", 0) for r in results)),
        ]
        for row_n, (label, value) in enumerate(summary, 1):
            cell_l = ws3.cell(row=row_n, column=1, value=label)
            cell_l.font = _LABEL_FONT
            cell_l.border = _THIN_BORDER
            cell_v = ws3.cell(row=row_n, column=2, value=value)
            cell_v.border = _THIN_BORDER
        ws3.column_dimensions["A"].width = 25
        ws3.column_dimensions["B"].width = 30

        wb.save(filepath)
        logger.log(f"Результаты сохранены: {filepath} ({total} аккаунтов)", "success")
        return filepath

    except Exception as e:
        logger.log(f"Ошибка экспорта XLSX: {e}", "error")
        return None
