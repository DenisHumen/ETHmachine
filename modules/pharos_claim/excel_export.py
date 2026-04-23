"""Экспорт результатов Pharos Claim Checker в XLSX (result/pharos_claim/)."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from modules.pharos_claim import database as db
from modules.simple_logger import logger as _logger

RESULT_DIR = Path(__file__).parent.parent.parent / "result" / "pharos_claim"

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _mask_proxy(p: Optional[str]) -> str:
    if not p:
        return ""
    return p.split("@", 1)[1] if "@" in p else p


def _auto_width(ws, n_cols: int, n_rows: int, max_w: int = 60) -> None:
    for col in range(1, n_cols + 1):
        max_len = 0
        for row in range(1, n_rows + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), max_w))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 3


def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGN
        c.border = _BORDER


def export_run(run_id: int, filename: Optional[str] = None) -> Optional[str]:
    tasks = db.get_run_tasks(run_id)
    if not tasks:
        _logger.warning(f"Нет задач в run_id={run_id}")
        return None

    run_info = db.get_run(run_id) or {}

    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    if filename is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pharos_claim_{run_id}_{ts}.xlsx"
    filepath = str(RESULT_DIR / filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "№", "Account", "Address", "Status", "Eligible",
        "Amount", "Claimed", "Proxy", "Endpoint", "Error", "Finished",
    ]
    _write_header(ws, headers)

    for i, t in enumerate(tasks, 1):
        row = i + 1
        status = t.get("status") or "unknown"
        eligible_val = t.get("eligible")
        eligible_str = ""
        if status == "completed":
            eligible_str = "YES" if eligible_val == 1 else "NO"

        claimed_val = t.get("claimed")
        claimed_str = "" if claimed_val is None else ("YES" if claimed_val else "NO")

        values = [
            i,
            t.get("account_name") or "",
            t.get("address") or "",
            status,
            eligible_str,
            t.get("amount") or "",
            claimed_str,
            _mask_proxy(t.get("proxy")),
            t.get("endpoint") or "",
            t.get("error") or "",
            t.get("finished_at") or "",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = _BORDER

        # подсветка Status / Eligible
        st_cell = ws.cell(row=row, column=4)
        if status == "completed":
            st_cell.fill = _OK_FILL if eligible_val == 1 else _WARN_FILL
        elif status == "failed":
            st_cell.fill = _FAIL_FILL
        st_cell.alignment = Alignment(horizontal="center")

        el_cell = ws.cell(row=row, column=5)
        if eligible_str == "YES":
            el_cell.fill = _OK_FILL
        elif eligible_str == "NO":
            el_cell.fill = _WARN_FILL
        el_cell.alignment = Alignment(horizontal="center")

    _auto_width(ws, len(headers), len(tasks) + 1)

    # Summary
    ws2 = wb.create_sheet("Summary")
    summary_rows = [
        ("Run ID", run_info.get("id", run_id)),
        ("Started",  run_info.get("started_at", "")),
        ("Finished", run_info.get("finished_at", "")),
        ("Total",    run_info.get("total", len(tasks))),
        ("Eligible", run_info.get("eligible", 0)),
        ("Not eligible", run_info.get("not_eligible", 0)),
        ("Failed",   run_info.get("failed", 0)),
        ("Exported", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    bold = Font(bold=True)
    for r, (label, value) in enumerate(summary_rows, 1):
        a = ws2.cell(row=r, column=1, value=label)
        b = ws2.cell(row=r, column=2, value=value)
        a.font = bold
        a.border = _BORDER
        b.border = _BORDER
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 40

    try:
        wb.save(filepath)
    except Exception as e:
        _logger.error(f"Ошибка сохранения XLSX: {e}")
        return None

    _logger.success(f"Результаты сохранены: {filepath} ({len(tasks)} кошельков)")
    return filepath


def export_latest_run() -> Optional[str]:
    run_id = db.get_last_run_id()
    if run_id is None:
        _logger.warning("В БД нет ни одного запуска")
        return None
    return export_run(run_id)
