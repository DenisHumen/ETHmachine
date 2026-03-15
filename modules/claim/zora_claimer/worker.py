"""
Zora Claimer Worker
Мультипоточная обработка клейма наград
"""

import os
import random
import time
from pathlib import Path
from datetime import datetime
from typing import List, Tuple, Dict, Optional
from concurrent.futures import ThreadPoolExecutor
from threading import Lock

from colorama import Fore, Style
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from modules.simple_logger import logger, setup_file_logging
from modules.proxy_manager import ProxyManager, mask_proxy
from config.modules.cfg_base import (
    NUM_THREADS, SLEEP_BETWEEN_ACTIONS, DELAY_BETWEEN_ACCOUNTS, RETRY_COUNT
)
from config.modules.cfg_claimer import (
    CLAIMER_NETWORKS, CLAIMER_USE_PROXY, CLAIMER_RANDOM_PROXY,
    CLAIMER_SHUFFLE_WALLETS, MIN_CLAIM_AMOUNT_ETH,
)
from modules.claim.zora_claimer.client import ZoraClaimerClient
from modules.claim.zora_claimer.database import (
    init_database, create_claim_tasks, get_pending_tasks, get_task,
    update_task_claimable, update_task_status, get_task_statistics,
    get_all_results, all_tasks_completed, reset_database, get_total_tasks_count,
)

project_root = Path(__file__).parent.parent.parent.parent
progress_lock = Lock()

# Статистика текущего запуска
run_stats = {
    'total_claimed_eth': 0.0,
    'total_gas_eth': 0.0,
    'successful': 0,
    'failed': 0,
    'skipped': 0,
}


def load_private_keys() -> List[str]:
    """Загрузить приватные ключи из файла"""
    keys_file = project_root / 'data' / 'private_keys.txt'

    if not keys_file.exists():
        logger.error(f"Файл {keys_file} не найден!")
        return []

    with open(keys_file, 'r') as f:
        keys = [line.strip() for line in f if line.strip() and not line.startswith('#')]

    logger.info(f"Загружено {len(keys)} приватных ключей")
    return keys


def prepare_wallets(private_keys: List[str]) -> List[Tuple[str, str]]:
    """Подготовить кошельки (address, private_key)"""
    from eth_account import Account
    wallets = []
    for pk in private_keys:
        try:
            pk_hex = pk if pk.startswith('0x') else f'0x{pk}'
            address = Account.from_key(pk_hex).address
            wallets.append((address, pk))
        except Exception as e:
            logger.error(f"Ошибка обработки ключа: {e}")
    return wallets


def get_enabled_networks() -> List[str]:
    """Получить список включённых сетей"""
    return [name for name, cfg in CLAIMER_NETWORKS.items() if cfg.get('enabled')]


def process_wallet_claim(
    private_key: str,
    wallet_address: str,
    proxy: Optional[str],
    all_proxies: List[str],
    wallet_index: int,
    total_wallets: int,
) -> Dict:
    """
    Обработка клейма для одного кошелька на всех сетях

    Returns:
        dict с результатами по каждой сети
    """
    results = {}
    networks = get_enabled_networks()
    proxy_display = mask_proxy(proxy) if proxy else 'No proxy'

    logger.info(
        f"[{wallet_index}/{total_wallets}] | [{wallet_address}] | "
        f"Proxy: {proxy_display} | Сети: {', '.join(networks)}"
    )

    client = ZoraClaimerClient(private_key=private_key, proxy=proxy)
    current_proxy = proxy

    for network in networks:
        net_name = CLAIMER_NETWORKS[network]['name']

        # Проверяем статус в БД
        task = get_task(wallet_address, network)
        if task and task.get('status') in ('completed', 'skipped'):
            logger.info(f"[{wallet_address}] {net_name} уже обработан, пропускаем")
            results[network] = {'status': 'skipped', 'reason': 'already_done'}
            continue

        update_task_status(wallet_address, network, 'in_progress')

        # Проверяем claimable баланс
        claimable_wei, claimable_eth = client.check_claimable(network)
        update_task_claimable(wallet_address, network, str(claimable_wei), claimable_eth)

        if claimable_eth < MIN_CLAIM_AMOUNT_ETH:
            msg = f'Нет доступного клейма ({claimable_eth:.8f} ETH)'
            logger.info(f"[{wallet_address}] {net_name}: {msg}")
            update_task_status(wallet_address, network, 'skipped', error_message=msg)
            results[network] = {'status': 'skipped', 'reason': msg}

            with progress_lock:
                run_stats['skipped'] += 1
            continue

        logger.info(
            f"[{wallet_address}] {net_name}: доступно {claimable_eth:.8f} ETH для клейма"
        )

        # Пробуем клеймить с ретраями
        success = False
        last_error = ""

        for attempt in range(RETRY_COUNT):
            # Обновляем клиент с новым прокси при ретрае
            if attempt > 0 and all_proxies and len(all_proxies) > 1:
                current_proxy = random.choice([p for p in all_proxies if p != current_proxy] or all_proxies)
                client = ZoraClaimerClient(private_key=private_key, proxy=current_proxy)
                logger.info(f"[{wallet_address}] Смена прокси, попытка {attempt + 1}/{RETRY_COUNT}")

            claim_result = client.claim(network)

            if claim_result['success']:
                logger.success(
                    f"[{wallet_address}] {net_name}: заклеймлено {claim_result['claimed_eth']:.8f} ETH | "
                    f"Газ: {claim_result['gas_cost_eth']:.8f} ETH | "
                    f"TX: {claim_result['explorer_link']}"
                )

                update_task_status(
                    wallet_address, network, 'completed',
                    tx_hash=claim_result['tx_hash'],
                    explorer_link=claim_result['explorer_link'],
                    claimed_wei=claim_result['claimed_wei'],
                    claimed_eth=claim_result['claimed_eth'],
                    gas_used=claim_result['gas_used'],
                    gas_price=claim_result['gas_price'],
                    gas_cost_eth=claim_result['gas_cost_eth'],
                )

                with progress_lock:
                    run_stats['successful'] += 1
                    run_stats['total_claimed_eth'] += claim_result['claimed_eth']
                    run_stats['total_gas_eth'] += claim_result['gas_cost_eth']

                results[network] = {'status': 'completed', 'result': claim_result}
                success = True
                break
            else:
                last_error = claim_result.get('error', 'Unknown error')
                logger.warning(
                    f"[{wallet_address}] {net_name}: {last_error} (попытка {attempt + 1}/{RETRY_COUNT})"
                )

                if attempt < RETRY_COUNT - 1:
                    time.sleep(random.uniform(2, 5))

        if not success:
            update_task_status(wallet_address, network, 'failed', error_message=last_error)
            results[network] = {'status': 'failed', 'error': last_error}

            with progress_lock:
                run_stats['failed'] += 1

        # Пауза между сетями
        if network != networks[-1]:
            delay = random.uniform(*SLEEP_BETWEEN_ACTIONS)
            time.sleep(delay)

    return results


def run_claimer():
    """Запуск клеймера с мультипоточностью"""
    global run_stats
    run_stats = {'total_claimed_eth': 0.0, 'total_gas_eth': 0.0, 'successful': 0, 'failed': 0, 'skipped': 0}

    # Настройка логирования в файл
    log_dir = project_root / 'log'
    log_dir.mkdir(exist_ok=True)
    setup_file_logging(str(log_dir / 'zora_claimer.log'))

    networks = get_enabled_networks()
    if not networks:
        logger.error("Нет включённых сетей в конфиге CLAIMER_NETWORKS!")
        return

    # Загружаем ключи и прокси
    private_keys = load_private_keys()
    if not private_keys:
        logger.error("Нет приватных ключей для обработки!")
        return

    wallets = prepare_wallets(private_keys)
    if not wallets:
        logger.error("Не удалось подготовить кошельки!")
        return

    proxies = ProxyManager.load_proxies() if CLAIMER_USE_PROXY else []

    if CLAIMER_SHUFFLE_WALLETS:
        random.shuffle(wallets)
        logger.info("Список кошельков перемешан")

    # Инициализация БД
    init_database()

    # Проверяем состояние БД
    if all_tasks_completed():
        logger.info("Все задачи завершены, очищаем БД для нового запуска...")
        deleted = reset_database()
        logger.success(f"Очищено {deleted} задач из БД")

    # Создаём задачи
    create_claim_tasks(wallets, networks)

    # Фильтруем кошельки с незавершёнными задачами
    wallets_with_pending = []
    for addr, pk in wallets:
        for network in networks:
            task = get_task(addr, network)
            if not task or task.get('status') not in ('completed', 'skipped'):
                wallets_with_pending.append((addr, pk))
                break

    if not wallets_with_pending:
        logger.success("Все задачи для всех кошельков уже выполнены!")
        return

    logger.info(f"Кошельков: {len(wallets_with_pending)}, Сетей: {len(networks)}, Потоков: {NUM_THREADS}")
    logger.info(f"Сети: {', '.join(networks)} | Мин. клейм: {MIN_CLAIM_AMOUNT_ETH} ETH")

    # Обработка в пуле потоков
    with ThreadPoolExecutor(max_workers=NUM_THREADS) as executor:
        futures_dict = {}
        proxy_index = 0

        for idx, (wallet_address, private_key) in enumerate(wallets_with_pending):
            proxy = None
            if proxies:
                if CLAIMER_RANDOM_PROXY:
                    proxy = random.choice(proxies)
                else:
                    proxy = proxies[proxy_index % len(proxies)]
                    proxy_index += 1

            future = executor.submit(
                process_wallet_claim,
                private_key, wallet_address, proxy, proxies,
                idx + 1, len(wallets_with_pending)
            )
            futures_dict[future] = wallet_address

            # Задержка между запусками потоков
            if idx < len(wallets_with_pending) - 1:
                delay = random.uniform(*DELAY_BETWEEN_ACCOUNTS)
                time.sleep(delay)

        # Собираем результаты
        for future in futures_dict:
            wallet_addr = futures_dict[future]
            try:
                future.result()
            except Exception as e:
                logger.error(f"[{wallet_addr}] Критическая ошибка: {e}")

    # Выводим статистику
    print_run_statistics()

    # Сохраняем результаты в Excel
    export_results_xlsx()


def print_run_statistics():
    """Вывод статистики текущего запуска"""
    logger.info("=" * 60)
    logger.info("СТАТИСТИКА ЗАПУСКА ZORA CLAIMER:")
    logger.info(f"   Успешных клеймов: {run_stats['successful']}")
    logger.info(f"   Неудачных: {run_stats['failed']}")
    logger.info(f"   Пропущено: {run_stats['skipped']}")
    logger.info(f"   Заклеймлено ETH: {run_stats['total_claimed_eth']:.8f}")
    logger.info(f"   Газ потрачено: {run_stats['total_gas_eth']:.8f} ETH")
    net_profit = run_stats['total_claimed_eth'] - run_stats['total_gas_eth']
    logger.info(f"   Чистая прибыль: {net_profit:.8f} ETH")
    logger.info("=" * 60)

    # Статистика из БД
    db_stats = get_task_statistics()
    if db_stats:
        print(f"\n{Fore.CYAN}Статистика из БД:{Style.RESET_ALL}")
        for network, statuses in db_stats.items():
            net_name = CLAIMER_NETWORKS.get(network, {}).get('name', network)
            print(f"\n  {Fore.CYAN}{net_name}:{Style.RESET_ALL}")
            for status, data in statuses.items():
                color = Fore.GREEN if status == 'completed' else (Fore.RED if status == 'failed' else Fore.YELLOW)
                print(
                    f"    {color}{status}: {data['count']} | "
                    f"Заклеймлено: {data['total_claimed']:.8f} ETH | "
                    f"Газ: {data['total_gas']:.8f} ETH{Style.RESET_ALL}"
                )


def export_results_xlsx(output_path: Optional[str] = None):
    """Экспорт результатов в Excel файл"""
    results = get_all_results()
    if not results:
        logger.warning("Нет данных для экспорта")
        return None

    result_dir = project_root / 'result'
    result_dir.mkdir(exist_ok=True)

    if not output_path:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = str(result_dir / f'zora_claimer_{timestamp}.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "Zora Claimer Results"

    # Заголовки
    headers = [
        '#', 'Wallet', 'Network', 'Status',
        'Claimable ETH', 'Claimed ETH', 'Gas Cost ETH',
        'TX Hash', 'Error', 'Completed At'
    ]

    # Стили
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(name='Consolas', bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Записываем данные
    status_colors = {
        'completed': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
        'failed': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
        'skipped': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
        'pending': PatternFill(start_color='D6D8DB', end_color='D6D8DB', fill_type='solid'),
    }

    for idx, row_data in enumerate(results, 1):
        row_num = idx + 1
        values = [
            idx,
            row_data['wallet_address'],
            row_data['network'],
            row_data['status'],
            row_data.get('claimable_eth', 0.0),
            row_data.get('claimed_eth', 0.0),
            row_data.get('gas_cost_eth', 0.0),
            row_data.get('tx_hash', ''),
            row_data.get('error_message', ''),
            row_data.get('completed_at', ''),
        ]

        status_fill = status_colors.get(row_data['status'], status_colors['pending'])

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            if col == 2:  # Wallet
                cell.font = Font(name='Consolas', size=10)
            elif col == 4:  # Status
                cell.fill = status_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col in (5, 6, 7):  # ETH values
                cell.number_format = '0.00000000'

    # Ширина колонок
    col_widths = [5, 44, 10, 12, 16, 16, 16, 68, 40, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = width

    # Итоговая строка
    total_row = len(results) + 2
    ws.cell(row=total_row, column=1, value='ИТОГО').font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(r.get('claimable_eth', 0) or 0 for r in results)).number_format = '0.00000000'
    ws.cell(row=total_row, column=6, value=sum(r.get('claimed_eth', 0) or 0 for r in results)).number_format = '0.00000000'
    ws.cell(row=total_row, column=7, value=sum(r.get('gas_cost_eth', 0) or 0 for r in results)).number_format = '0.00000000'

    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).border = thin_border

    try:
        wb.save(output_path)
        logger.success(f"Результаты сохранены: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Ошибка сохранения Excel: {e}")
        return None
