"""
DeBank Protocol Checker
Проверка DeFi-позиций (стейкинг, лендинг, locked, LP и т.д.) через DeBank
Использует Playwright для перехвата portfolio/project_list API
"""

import re
import sys
import json
import math
import asyncio
import time
import random
from pathlib import Path
from datetime import datetime
from urllib.parse import urlparse

from rich.console import Console
from rich.live import Live
from rich.panel import Panel
from rich.text import Text
from rich.table import Table
from rich.console import Group
from questionary import Choice, select

project_root = Path(__file__).parent.parent.parent
sys.path.append(str(project_root))

from config.modules.general_config import (
    NUM_THREADS, RETRY_COUNT, SLEEP_BETWEEN_ACTIONS, DELAY_BETWEEN_ACCOUNTS
)
from config.modules.cfg_debank_protocol import (
    MIN_VALUE_USD, GRADIENT_MIN_USD, GRADIENT_MAX_USD
)
from modules.debank.database import (
    init_database, create_protocol_tasks, get_pending_protocol_tasks,
    update_protocol_task_status, save_protocol_positions_batch,
    delete_wallet_protocols, get_all_protocols, reset_protocols_database,
    get_protocol_task_statistics
)
from modules.debank.debank_checker import (
    load_wallets, load_private_keys_as_wallets, load_proxies, parse_proxy_for_playwright,
    load_wallet_rows, SOURCE_PRIVATE_KEYS, chains_from_body, USED_CHAINS_PATH,
    DATA_WAIT_TIMEOUT, MAX_ATTEMPTS
)

console = Console()
MAX_CONCURRENT = NUM_THREADS


_ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


def _sanitize_for_xlsx(text: str) -> str:
    """Удаляет символы запрещённые в xlsx ячейках"""
    return _ILLEGAL_CHARS_RE.sub('', text) if text else ''


def _safe_float(val, default=0.0) -> float:
    """Безопасная конвертация в float"""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _fmt_balance(val: float) -> str:
    """Форматирование баланса без научной нотации"""
    if val == 0:
        return '0'
    if val >= 1:
        return f"{val:.4f}"
    elif val >= 0.0001:
        return f"{val:.8f}"
    else:
        return f"{val:.12f}"


def _fmt_usd(val: float) -> str:
    """Форматирование USD"""
    if val == 0:
        return '$0.00'
    if val >= 0.01:
        return f"${val:.2f}"
    elif val >= 0.0001:
        return f"${val:.6f}"
    else:
        return f"${val:.10f}"


def _parse_token_info(token: dict) -> dict | None:
    """Парсинг одного токена из API — возвращает чистые данные или None"""
    if not isinstance(token, dict):
        return None
    amount = _safe_float(token.get('amount', 0))
    if amount <= 0:
        return None
    symbol = token.get('symbol') or token.get('optimized_symbol') or 'UNKNOWN'
    price = _safe_float(token.get('price', 0))
    return {
        'symbol': symbol,
        'amount': amount,
        'price': price,
        'value_usd': amount * price,
    }


def parse_protocol_data(data) -> list:
    """Парсинг DeFi-позиций из API ответа portfolio/project_list"""
    positions = []

    if isinstance(data, dict) and 'data' in data:
        protocol_list = data['data']
    elif isinstance(data, list):
        protocol_list = data
    else:
        return positions

    if not isinstance(protocol_list, list):
        return positions

    for protocol in protocol_list:
        if not isinstance(protocol, dict):
            continue

        protocol_name = protocol.get('name') or 'Unknown'
        protocol_id = protocol.get('id') or protocol.get('dao_id') or ''
        chain = protocol.get('chain') or ''

        portfolio_items = protocol.get('portfolio_item_list') or []
        if not isinstance(portfolio_items, list):
            continue

        for item in portfolio_items:
            if not isinstance(item, dict):
                continue

            position_type = item.get('name') or 'Unknown'
            stats = item.get('stats') or {}
            detail = item.get('detail') or {}
            net_usd = _safe_float(stats.get('net_usd_value', 0))
            asset_usd = _safe_float(stats.get('asset_usd_value', 0))

            description = detail.get('description') or ''

            # Unlock time
            unlock_time = ''
            for ts_key in ('unlock_at', 'end_at'):
                if ts_key in detail and detail[ts_key]:
                    try:
                        unlock_time = datetime.fromtimestamp(detail[ts_key]).strftime('%Y/%m/%d %H:%M')
                    except Exception:
                        pass
                    break

            # Health rate
            health_rate = ''
            hr = detail.get('health_rate')
            if hr is not None and isinstance(hr, (int, float)):
                if hr > 1e10:
                    health_rate = '>10'
                elif hr > 0:
                    health_rate = f"{hr:.2f}"

            # Токены
            supply_tokens = detail.get('supply_token_list') or []
            reward_tokens = detail.get('reward_token_list') or []
            borrow_tokens = detail.get('borrow_token_list') or []

            # Парсим reward/borrow в читаемый текст
            rewards_text = ''
            parsed_rewards = [_parse_token_info(r) for r in reward_tokens]
            parsed_rewards = [r for r in parsed_rewards if r]
            if parsed_rewards:
                parts = [f"{_fmt_balance(r['amount'])} {r['symbol']} ({_fmt_usd(r['value_usd'])})"
                         for r in parsed_rewards]
                rewards_text = 'Rewards: ' + ', '.join(parts)

            borrows_text = ''
            parsed_borrows = [_parse_token_info(r) for r in borrow_tokens]
            parsed_borrows = [r for r in parsed_borrows if r]
            if parsed_borrows:
                parts = [f"{_fmt_balance(r['amount'])} {r['symbol']} ({_fmt_usd(r['value_usd'])})"
                         for r in parsed_borrows]
                borrows_text = 'Borrows: ' + ', '.join(parts)

            extra_text = ' | '.join(filter(None, [rewards_text, borrows_text]))

            for token in supply_tokens:
                info = _parse_token_info(token)
                if not info:
                    continue

                balance_token = f"{_fmt_balance(info['amount'])} {info['symbol']}"
                pool_name = description if description else info['symbol']

                positions.append({
                    'protocol_name': protocol_name,
                    'protocol_id': protocol_id,
                    'chain': chain,
                    'position_type': position_type,
                    'pool_name': pool_name,
                    'balance': info['amount'],
                    'balance_token': balance_token,
                    'value_usd': info['value_usd'],
                    'unlock_time': unlock_time,
                    'health_rate': health_rate,
                    'extra_data': extra_text,
                })

    return positions


PROJECT_LIST_PATH = '/portfolio/project_list'


def watch_protocol_api(page) -> dict:
    """Подписаться на ответы DeBank; словарь наполняется по ходу загрузки страницы."""
    state = {'chains': None, 'projects': None}

    async def handle_response(response):
        url = response.url
        if 'api.debank.com' not in url:
            return
        path = urlparse(url).path
        try:
            if path == USED_CHAINS_PATH:
                chains = chains_from_body(await response.json())
                if chains is not None:
                    state['chains'] = chains
            elif path == PROJECT_LIST_PATH:
                state['projects'] = await response.json()
        except Exception:
            # Тело ответа могло стать недоступным — дождёмся следующего запроса.
            pass

    page.on('response', handle_response)
    return state


def protocols_ready(state: dict) -> bool:
    """Пришло ли всё, что DeBank собирался отдать по этому кошельку."""
    chains = state.get('chains')
    if chains is None:
        return False        # профиль ещё не загрузился
    if not chains:
        return True         # сетей нет — кошелёк пустой, позиций не будет
    return state.get('projects') is not None


async def wait_for_protocols(state: dict, timeout: float = DATA_WAIT_TIMEOUT) -> bool:
    """Ждём сами данные, а не фиксированную паузу."""
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if protocols_ready(state):
            return True
        await asyncio.sleep(0.5)
    return protocols_ready(state)


async def check_wallet_protocols(wallet: str, proxy_config: dict, semaphore: asyncio.Semaphore,
                                 playwright_instance) -> dict:
    """Проверка DeFi-позиций одного кошелька через Playwright"""
    async with semaphore:
        last_error = 'Max retries exceeded'

        for attempt in range(MAX_ATTEMPTS):
            try:
                launch_args = {'headless': True}
                if proxy_config:
                    launch_args['proxy'] = proxy_config

                browser = await playwright_instance.chromium.launch(**launch_args)
                try:
                    page = await browser.new_page()
                    state = watch_protocol_api(page)

                    # Не networkidle: профиль тянет данные десятками запросов и
                    # тишины в сети может не наступить вовсе — ждём сами данные.
                    await page.goto(
                        f'https://debank.com/profile/{wallet}',
                        wait_until='domcontentloaded',
                        timeout=45000
                    )

                    await wait_for_protocols(state)

                    is_last = attempt == MAX_ATTEMPTS - 1

                    if state['chains'] is None:
                        # Нет даже списка сетей — блокировка, капча или мёртвый прокси.
                        last_error = 'DeBank не отдал данные профиля'
                    elif protocols_ready(state) or is_last:
                        # Пустой кошелёк — тоже успех: сетей нет, значит и позиций нет.
                        positions = parse_protocol_data(state['projects']) if state['projects'] else []
                        return {
                            'wallet': wallet,
                            'positions': positions,
                            'success': True,
                            'total_usd': sum(p['value_usd'] for p in positions),
                            'error': None,
                        }
                    else:
                        last_error = 'Позиции не пришли за отведённое время'

                finally:
                    await browser.close()

            except Exception as e:
                last_error = str(e)[:80]

            if attempt < MAX_ATTEMPTS - 1:
                await asyncio.sleep(random.uniform(*SLEEP_BETWEEN_ACTIONS))

    return {'wallet': wallet, 'positions': [], 'success': False,
            'total_usd': 0, 'error': last_error}


def create_protocol_panel(total: int, success: int, failed: int, logs: list) -> Panel:
    """Панель прогресса для Protocol Checker"""
    processed = success + failed
    percent = (processed / total * 100) if total > 0 else 0

    bar_width = 40
    filled = int(bar_width * percent / 100)
    bar_filled = "━" * filled
    bar_empty = "─" * (bar_width - filled)

    success_rate = (success / processed * 100) if processed > 0 else 100
    bar_style = "bright_green" if success_rate >= 80 else "yellow" if success_rate >= 50 else "red"

    table = Table(show_header=False, box=None, padding=(0, 1))
    table.add_column("L1", style="dim")
    table.add_column("V1")
    table.add_column("L2", style="dim")
    table.add_column("V2")
    table.add_row(
        "✅ Успешно:", Text(str(success), style="bold green"),
        "❌ Ошибки:", Text(str(failed), style="bold red")
    )
    table.add_row(
        "📊 Всего:", Text(str(total), style="bold"),
        "⚡ Параллельно:", Text(str(MAX_CONCURRENT), style="bold cyan")
    )

    header = Text()
    header.append("\n🔗 ", style="bold")
    header.append("DeBank Protocol Checker\n", style="bold magenta")

    progress_bar = Text()
    progress_bar.append(bar_filled, style=f"bold {bar_style}")
    progress_bar.append(bar_empty, style="dim")
    progress_bar.append(f" {percent:.1f}%\n", style=f"bold {bar_style}")

    processed_text = Text()
    processed_text.append("Обработано: ", style="dim")
    processed_text.append(f"{processed}/{total}\n\n", style="bold")

    logs_section = Text()
    logs_section.append("📝 Последние события:\n", style="bold")
    for ts, msg, lvl in logs[-8:]:
        logs_section.append(f"  {ts} ", style="dim")
        style = {"SUCCESS": "green", "ERROR": "red", "WARNING": "yellow", "INFO": "cyan"}.get(lvl, "white")
        logs_section.append(f"{msg}\n", style=style)
    if not logs:
        logs_section.append("  Ожидание...\n", style="dim")

    content = Group(header, progress_bar, processed_text, table, Text(""), logs_section)

    return Panel(
        content,
        title="[bold bright_magenta]🔗 DEBANK PROTOCOL CHECKER[/bold bright_magenta]",
        subtitle="[dim]ETHmachine[/dim]",
        border_style="bright_magenta",
        padding=(1, 2)
    )


async def process_wallets_protocols_async(wallets: list, proxy_map: dict = None) -> dict:
    """Асинхронная обработка кошельков — сбор DeFi-позиций.

    ``proxy_map`` — прокси из строки самого кошелька в data.csv; для кошельков
    без своего прокси берётся общий пул по кругу.
    """
    from playwright.async_api import async_playwright

    results = {}
    total = len(wallets)
    success = 0
    failed = 0
    logs = []

    proxies = load_proxies()
    proxy_map = proxy_map or {}
    paired = sum(1 for w in wallets if proxy_map.get(w))
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)

    if not proxies:
        console.print("[yellow][!] Прокси не найдены, запросы пойдут напрямую[/yellow]")

    logs.append((time.strftime("%H:%M:%S"), f"Запуск {total} кошельков ({MAX_CONCURRENT} параллельно)", "INFO"))
    if paired:
        logs.append((time.strftime("%H:%M:%S"),
                     f"У {paired} кошельков свой прокси из data.csv", "INFO"))
    if proxies:
        logs.append((time.strftime("%H:%M:%S"),
                     f"Общий пул прокси: {len(proxies)} (round-robin)", "INFO"))

    live = Live(create_protocol_panel(total, 0, 0, logs), console=console, refresh_per_second=2)
    live.start()

    try:
        async with async_playwright() as p:

            async def process_wallet(idx, wallet):
                nonlocal success, failed
                # Свой прокси кошелька важнее общего пула: один адрес — один IP.
                proxy_str = proxy_map.get(wallet)
                if not proxy_str and proxies:
                    proxy_str = proxies[idx % len(proxies)]
                proxy_config = parse_proxy_for_playwright(proxy_str)

                short = f"{wallet[:6]}...{wallet[-4:]}"
                try:
                    result = await check_wallet_protocols(wallet, proxy_config, semaphore, p)
                    results[wallet] = result

                    if result['success']:
                        success += 1
                        delete_wallet_protocols(wallet)
                        if result['positions']:
                            batch = [
                                (wallet, pos['protocol_name'], pos['protocol_id'], pos['chain'],
                                 pos['position_type'], pos['pool_name'], pos['balance'],
                                 pos['balance_token'], pos['value_usd'], pos['unlock_time'],
                                 pos['health_rate'], pos['extra_data'])
                                for pos in result['positions']
                            ]
                            save_protocol_positions_batch(batch)

                        update_protocol_task_status(wallet, 'completed')
                        pos_count = len(result['positions'])
                        total_usd = result.get('total_usd', 0)
                        logs.append((
                            time.strftime("%H:%M:%S"),
                            f"[{short}] ✅ {pos_count} позиций | ${total_usd:.2f}",
                            "SUCCESS"
                        ))
                    else:
                        failed += 1
                        update_protocol_task_status(wallet, 'failed', result.get('error', 'Unknown'))
                        logs.append((
                            time.strftime("%H:%M:%S"),
                            f"[{short}] ❌ {result.get('error', 'Unknown')[:40]}",
                            "ERROR"
                        ))
                except Exception as e:
                    failed += 1
                    results[wallet] = {'wallet': wallet, 'positions': [], 'success': False, 'error': str(e)[:50]}
                    update_protocol_task_status(wallet, 'failed', str(e)[:100])
                    logs.append((time.strftime("%H:%M:%S"), f"[{short}] ❌ {str(e)[:30]}", "ERROR"))

                live.update(create_protocol_panel(total, success, failed, logs))

            # Очередь и пул воркеров вместо создания всех задач заранее: раньше
            # между стартами ждали DELAY_BETWEEN_ACCOUNTS последовательно, и на
            # тысячах кошельков одна только раздача задач растягивалась на часы,
            # а реальная параллельность была в разы ниже MAX_CONCURRENT.
            queue = asyncio.Queue()
            for item in enumerate(wallets):
                queue.put_nowait(item)

            async def worker():
                while True:
                    try:
                        idx, wallet = queue.get_nowait()
                    except asyncio.QueueEmpty:
                        return
                    await process_wallet(idx, wallet)
                    # Пауза между кошельками одного воркера, чтобы не бить DeBank пачкой.
                    await asyncio.sleep(random.uniform(*DELAY_BETWEEN_ACCOUNTS))

            workers = [asyncio.create_task(worker())
                       for _ in range(min(MAX_CONCURRENT, len(wallets)))]
            await asyncio.gather(*workers, return_exceptions=True)

    finally:
        live.stop()

    return results


def process_wallets_protocols(wallets: list, proxy_map: dict = None) -> dict:
    """Обёртка для запуска асинхронной обработки"""
    return asyncio.run(process_wallets_protocols_async(wallets, proxy_map))


# ─── XLSX Export ─────────────────────────────────────────────────────────────


def _gradient_color(value_usd: float) -> str | None:
    """Возвращает hex-цвет градиента от белого к зелёному в зависимости от суммы.
    None — без заливки (ниже порога GRADIENT_MIN_USD).
    """
    if value_usd < GRADIENT_MIN_USD:
        return None

    # Логарифмический градиент для лучшего распределения цветов
    if value_usd >= GRADIENT_MAX_USD:
        t = 1.0
    else:
        log_min = math.log10(max(GRADIENT_MIN_USD, 0.01))
        log_max = math.log10(max(GRADIENT_MAX_USD, 1.0))
        log_val = math.log10(max(value_usd, GRADIENT_MIN_USD))
        t = (log_val - log_min) / (log_max - log_min)
        t = max(0.0, min(1.0, t))

    # Градиент: белый (FFFFFF) → светло-зелёный (C6EFCE) → ярко-зелёный (27AE60)
    if t <= 0.5:
        # белый → светло-зелёный
        s = t * 2
        r = int(255 - (255 - 198) * s)
        g = int(255 - (255 - 239) * s)
        b = int(255 - (255 - 206) * s)
    else:
        # светло-зелёный → ярко-зелёный
        s = (t - 0.5) * 2
        r = int(198 - (198 - 39) * s)
        g = int(239 - (239 - 174) * s)
        b = int(206 - (206 - 96) * s)

    return f"{r:02X}{g:02X}{b:02X}"


def _format_cell_text(p: dict) -> str:
    """Форматирование ячейки — чистый читаемый текст без JSON"""
    parts = [p.get('balance_token', ''), _fmt_usd(p.get('value_usd', 0))]

    if p.get('unlock_time'):
        parts.append(f"unlock: {p['unlock_time']}")
    if p.get('health_rate'):
        parts.append(f"health: {p['health_rate']}")

    extra = p.get('extra_data', '')
    if extra and isinstance(extra, str) and not extra.startswith('{'):
        parts.append(extra)
    elif extra and isinstance(extra, str) and extra.startswith('{'):
        # Старые данные в JSON — декодируем в текст
        try:
            data = json.loads(extra)
            if isinstance(data, dict):
                for key in ('rewards', 'borrows'):
                    items = data.get(key, [])
                    for item in items:
                        sym = item.get('symbol', '?')
                        amt = _safe_float(item.get('amount', 0))
                        val = _safe_float(item.get('value_usd', 0))
                        parts.append(f"{key[:-1]}: {_fmt_balance(amt)} {sym} ({_fmt_usd(val)})")
        except (json.JSONDecodeError, TypeError):
            pass

    return ' | '.join(filter(None, parts))


def _make_column_key(p: dict) -> str:
    """Уникальный ключ колонки"""
    return f"{p['protocol_name']}|{p['chain']}|{p['position_type']}|{p['pool_name']}"


def _make_column_header(key: str) -> str:
    """Читаемый заголовок колонки"""
    parts = key.split('|')
    if len(parts) >= 4:
        return f"{parts[0]} [{parts[1]}] {parts[2]}: {parts[3]}"
    return key


def save_protocols_xlsx(wallets: list):
    """Экспорт DeFi-позиций в XLSX с цветным градиентом по суммам"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    result_dir = project_root / 'result'
    result_dir.mkdir(exist_ok=True)

    all_protocols = get_all_protocols()
    if not all_protocols:
        console.print("[yellow][!] Нет данных о протоколах для экспорта[/yellow]")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = result_dir / f"debank_protocols_{timestamp}.xlsx"

    # Собираем данные: wallet → {col_key: {text, value_usd}}
    wallet_data = {}
    all_columns = []
    seen_columns = set()

    for p in all_protocols:
        value_usd = _safe_float(p.get('value_usd', 0))

        # Фильтр по минимальной сумме
        if value_usd < MIN_VALUE_USD:
            continue

        w = p['wallet_address']
        col_key = _make_column_key(p)

        if col_key not in seen_columns:
            seen_columns.add(col_key)
            all_columns.append(col_key)

        wallet_data.setdefault(w, {})[col_key] = {
            'text': _format_cell_text(p),
            'value_usd': value_usd,
        }

    if not wallet_data:
        console.print(f"[yellow][!] Нет позиций >= ${MIN_VALUE_USD:.2f} для экспорта[/yellow]")
        return None

    # Порядок кошельков: сначала из файла, потом остальные
    all_wallets = list(dict.fromkeys(
        [w for w in wallets if w in wallet_data] +
        [w for w in wallet_data if w not in wallets]
    ))

    # Создаём XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "DeFi Protocols"

    # Стили
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC'),
    )
    wallet_font = Font(name='Consolas', size=10)
    cell_font = Font(size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)

    # Заголовки
    headers = ['#', 'Wallet', 'Total USD'] + [_make_column_header(c) for c in all_columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=_sanitize_for_xlsx(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Данные
    for row_idx, wallet in enumerate(all_wallets, 2):
        positions = wallet_data.get(wallet, {})
        total_usd = sum(pos['value_usd'] for pos in positions.values())

        # #
        cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        cell.font = cell_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

        # Wallet
        cell = ws.cell(row=row_idx, column=2, value=wallet)
        cell.font = wallet_font
        cell.border = thin_border

        # Total USD
        cell = ws.cell(row=row_idx, column=3, value=round(total_usd, 2))
        cell.font = Font(bold=True, size=10)
        cell.number_format = '$#,##0.00'
        cell.border = thin_border
        color = _gradient_color(total_usd)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Позиции
        for col_offset, col_key in enumerate(all_columns):
            cell_col = 4 + col_offset
            pos = positions.get(col_key)
            cell = ws.cell(row=row_idx, column=cell_col)
            cell.border = thin_border
            cell.font = cell_font
            cell.alignment = cell_alignment

            if pos:
                cell.value = _sanitize_for_xlsx(pos['text'])
                color = _gradient_color(pos['value_usd'])
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            else:
                cell.value = ''

    # Ширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 14
    for col_offset in range(len(all_columns)):
        col_letter = ws.cell(row=1, column=4 + col_offset).column_letter
        ws.column_dimensions[col_letter].width = 35

    # Закрепляем заголовок и первые 3 колонки
    ws.freeze_panes = 'D2'

    # Автофильтр
    ws.auto_filter.ref = ws.dimensions

    wb.save(str(filepath))
    console.print(f"[green][v] Протоколы сохранены: {filepath}[/green]")
    console.print(f"[dim]   Кошельков: {len(all_wallets)} | Протоколов: {len(all_columns)} | "
                  f"Фильтр: >= ${MIN_VALUE_USD:.2f} | Градиент от ${GRADIENT_MIN_USD:.2f}[/dim]")
    return filepath


def print_protocol_summary(results: dict):
    """Вывод итогов проверки протоколов"""
    success_list = [r for r in results.values() if r['success']]
    failed_list = [r for r in results.values() if not r['success']]

    total_positions = sum(len(r['positions']) for r in success_list)
    total_usd = sum(r.get('total_usd', 0) for r in success_list)

    protocols_found = set()
    for r in success_list:
        for pos in r['positions']:
            protocols_found.add(pos['protocol_name'])

    console.print("\n" + "=" * 60)
    console.print("[bold magenta]🔗 ИТОГИ ПРОВЕРКИ ПРОТОКОЛОВ[/bold magenta]")
    console.print("=" * 60)
    console.print(f"[green]✅ Успешно: {len(success_list)}[/green]")
    console.print(f"[red]❌ Ошибки: {len(failed_list)}[/red]")
    console.print(f"[cyan]🔗 Протоколов найдено: {len(protocols_found)}[/cyan]")
    console.print(f"[cyan]📋 Всего позиций: {total_positions}[/cyan]")
    console.print(f"[yellow]💰 Общая стоимость: ${total_usd:,.2f}[/yellow]")
    console.print("=" * 60)

    if protocols_found:
        console.print(f"\n[dim]Протоколы: {', '.join(sorted(protocols_found))}[/dim]")

    wallets_with_value = sorted(
        [r for r in success_list if r.get('total_usd', 0) > 0],
        key=lambda x: x.get('total_usd', 0),
        reverse=True
    )
    if wallets_with_value:
        console.print("\n[bold magenta]🏆 Топ кошельков по стоимости DeFi:[/bold magenta]")
        for i, r in enumerate(wallets_with_value[:10], 1):
            w = r['wallet']
            console.print(
                f"  {i}. {w[:10]}...{w[-6:]} → "
                f"[green]${r['total_usd']:,.2f}[/green] "
                f"({len(r['positions'])} позиций)"
            )


def debank_protocol_menu():
    """Главное меню DeBank Protocol Checker"""
    wallet_source = select(
        "\n╔════════════════════════════════════════════════╗\n"
        "║      Источник кошельков                        ║\n"
        "╚════════════════════════════════════════════════╝",
        choices=[
            Choice('   📋 Адреса кошельков (колонка wallet_address в data.csv)', 'wallets'),
            Choice('   🔑 Приватные ключи (колонка private_key в data.csv)', 'private_keys'),
            Choice('   🔙 Назад', 'back')
        ],
        qmark='🛠️ ',
        pointer='👉'
    ).ask()

    if wallet_source == 'back' or not wallet_source:
        return

    rows = load_wallet_rows(wallet_source)
    if not rows:
        column = 'private_key' if wallet_source == SOURCE_PRIVATE_KEYS else 'wallet_address'
        console.print(f"[red]❌ В data.csv не заполнена колонка {column}[/red]")
        return

    wallets = [row['address'] for row in rows]
    # Прокси кошелька из его же строки data.csv.
    proxy_map = {row['address']: row['proxy'] for row in rows if row['proxy']}

    if wallet_source == SOURCE_PRIVATE_KEYS:
        console.print(f"[cyan]🔑 Получено {len(wallets)} адресов из приватных ключей[/cyan]")
    else:
        console.print(f"[cyan]📋 Загружено {len(wallets)} кошельков[/cyan]")

    init_database()

    stats = get_protocol_task_statistics()
    completed = stats.get('completed', 0)
    pending = stats.get('pending', 0)
    failed_count = stats.get('failed', 0)

    if completed > 0 or pending > 0 or failed_count > 0:
        console.print(f"[dim]БД: ✅ {completed} | ⏳ {pending} | ❌ {failed_count}[/dim]")

    action = select(
        "\n╔════════════════════════════════════════════════╗\n"
        "║      DeBank Protocol Checker                   ║\n"
        "╚════════════════════════════════════════════════╝",
        choices=[
            Choice('   ▶️  Продолжить незавершённые задачи', 'continue'),
            Choice('   🔄 Начать заново (сброс БД)', 'reset'),
            Choice('   📊 Экспорт результатов в XLSX', 'export'),
            Choice('   🔙 Назад', 'back')
        ],
        qmark='🛠️ ',
        pointer='👉'
    ).ask()

    if action == 'back' or not action:
        return

    if action == 'export':
        save_protocols_xlsx(wallets)
        return

    if action == 'reset':
        reset_protocols_database()
        console.print("[yellow]🔄 База данных протоколов сброшена[/yellow]")
        create_protocol_tasks(wallets)
    else:
        create_protocol_tasks(wallets)

    pending_tasks = get_pending_protocol_tasks()
    if not pending_tasks:
        console.print("[yellow][!] Все задачи уже выполнены. Используйте 'Начать заново' для повторной проверки.[/yellow]")
        save_protocols_xlsx(wallets)
        return

    wallets_to_check = [t['wallet_address'] for t in pending_tasks]
    console.print(f"[cyan]📋 Задач к выполнению: {len(wallets_to_check)}[/cyan]")

    results = process_wallets_protocols(wallets_to_check, proxy_map)

    print_protocol_summary(results)

    save_protocols_xlsx(wallets)

    console.print("\n[bold green]✅ Проверка протоколов DeBank завершена![/bold green]\n")

    try:
        results_list = list(results.values())
        success_count = len([r for r in results_list if r['success']])
        total_usd = sum(r.get('total_usd', 0) for r in results_list if r['success'])
    except Exception:
        pass


if __name__ == "__main__":
    debank_protocol_menu()
