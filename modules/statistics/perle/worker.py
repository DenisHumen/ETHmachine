"""
Worker для Perle Eligibility Checker
Проверка элигибельности SOL кошельков для Perle airdrop
Эмулирует подключение кошелька Phantom через браузер
"""

import json
import time
import random
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Event

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from colorama import Fore, Style
from solders.keypair import Keypair
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from modules.simple_logger import logger, log_task, log_wallet_task
from modules.proxy_manager import get_proxy_dict, mask_proxy
from modules.data_manager import load_data
from modules.statistics.perle.database import (
    init_database, create_check_tasks, get_pending_tasks,
    update_task_success, update_task_failed,
    get_all_results, get_task_statistics,
)
from config.modules.cfg_base import (
    NUM_THREADS,
    SLEEP_BETWEEN_ACTIONS,
    DELAY_BETWEEN_ACCOUNTS,
    RETRY_COUNT,
)

PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
RESULT_DIR = PROJECT_ROOT / 'result' / 'statistics' / 'perle'
API_URL = 'https://register.perle.xyz/api/eligibility'

_stop_event = Event()

# ═══════════════════════════════════════════════════
# Phantom-like browser headers
# Эмуляция подключения кошелька Phantom через браузер
# ═══════════════════════════════════════════════════

_USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.2 Safari/605.1.15',
]


def _get_phantom_headers() -> dict:
    """Возвращает заголовки, эмулирующие запрос с сайта Perle через Phantom wallet"""
    ua = random.choice(_USER_AGENTS)
    return {
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
        'Content-Type': 'application/json',
        'Origin': 'https://register.perle.xyz',
        'Referer': 'https://register.perle.xyz/flow',
        'Sec-Ch-Ua': '"Chromium";v="131", "Not_A Brand";v="24"',
        'Sec-Ch-Ua-Mobile': '?0',
        'Sec-Ch-Ua-Platform': '"Windows"',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': ua,
    }


def _derive_sol_address(private_key: str) -> Optional[str]:
    """Получить SOL адрес из приватного ключа"""
    try:
        kp = Keypair.from_base58_string(private_key.strip())
        return str(kp.pubkey())
    except Exception as e:
        logger.error(f"Не удалось получить адрес из ключа: {e}")
        return None


def _create_session(proxy: Optional[str] = None) -> requests.Session:
    """Создать сессию с retry-адаптером для устойчивости к SSL ошибкам через прокси"""
    session = requests.Session()

    # Автоматический retry на транспортном уровне для SSL/Connection ошибок
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[502, 503, 504, 520, 521, 522, 523, 524],
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount('https://', adapter)
    session.mount('http://', adapter)

    if proxy:
        proxy_dict = get_proxy_dict(proxy)
        if proxy_dict:
            session.proxies.update(proxy_dict)

    return session


def _check_eligibility(sol_address: str, proxy: Optional[str] = None, timeout: int = 30) -> dict:
    """
    Проверить элигибельность кошелька через API.
    Эмулирует запрос как если бы Phantom wallet подключился на сайте.
    """
    headers = _get_phantom_headers()
    session = _create_session(proxy)

    response = session.post(
        API_URL,
        json={'wallet': sol_address},
        headers=headers,
        timeout=timeout,
    )
    response.raise_for_status()
    data = response.json()
    logger.debug(f"API response for {sol_address}: {json.dumps(data)}")
    return data


def _is_ssl_or_proxy_error(exc: Exception) -> bool:
    """Определяет, связана ли ошибка с SSL/прокси/таймаутом."""
    text = str(exc)
    return any(kw in text for kw in ('SSL', 'EOF', 'timed out', 'ProxyError'))


def _process_single_wallet(
    index: int,
    total: int,
    sol_private_key: str,
    proxy: Optional[str] = None,
    account_name: Optional[str] = None,
) -> dict:
    """Обработка одного кошелька с ретраями из конфига."""
    sol_address = _derive_sol_address(sol_private_key)
    if not sol_address:
        return {'address': 'UNKNOWN', 'error': 'Не удалось получить адрес из ключа'}

    proxy_info = f" │ proxy: {mask_proxy(proxy)}" if proxy else ""
    max_retries = max(RETRY_COUNT, 1)
    last_error = None

    for attempt in range(1, max_retries + 1):
        if _stop_event.is_set():
            return {'address': sol_address, 'error': 'Остановлено'}

        try:
            data = _check_eligibility(sol_address, proxy)
            eligible = data.get('eligible', False)
            total_amount = data.get('total', '0')
            details = json.dumps(data.get('details', {}))

            update_task_success(sol_address, eligible, total_amount, details)

            status = "success" if eligible else "warning"
            label  = f"✅ ELIGIBLE │ Total: {total_amount}" if eligible else "❌ NOT ELIGIBLE"
            log_wallet_task(sol_address, index, total, f"{label}{proxy_info}", status, account_name=account_name)

            return {
                'address': sol_address,
                'eligible': eligible,
                'total': total_amount,
                'details': details,
            }

        except requests.exceptions.RequestException as e:
            last_error = str(e)[:150]
            ssl_err = _is_ssl_or_proxy_error(e)
            base_delay = SLEEP_BETWEEN_ACTIONS
            if ssl_err:
                base_delay = [base_delay[0] * 2, base_delay[1] * 2]

            if attempt < max_retries:
                retry_delay = random.uniform(*base_delay)
                log_task(
                    index, total,
                    f"⚠️ Попытка {attempt}/{max_retries}: {last_error} │ Повтор через {retry_delay:.1f}с",
                    "warning",
                    account_name=account_name,
                )
                time.sleep(retry_delay)
            else:
                update_task_failed(sol_address, last_error)
                tag = '🔌 SSL/PROXY' if ssl_err else '❌ FAIL'
                log_wallet_task(
                    sol_address, index, total,
                    f"{tag} после {max_retries} попыток: {last_error}{proxy_info}", "error",
                    account_name=account_name,
                )

    return {'address': sol_address, 'error': last_error}


def _thread_worker(args: Tuple) -> dict:
    """Обёртка для потока: добавляет задержку между аккаунтами"""
    index, total, sol_pk, proxy, account_name = args

    # Задержка между стартом аккаунтов (кроме первого)
    if index > 1:
        delay = random.uniform(*DELAY_BETWEEN_ACCOUNTS)
        time.sleep(delay)

    return _process_single_wallet(index, total, sol_pk, proxy, account_name=account_name)


def run_checker():
    """Запуск чекера элигибельности с мультипоточностью"""
    _stop_event.clear()
    init_database()

    rows = load_data()
    if not rows:
        logger.error("Нет данных в data.csv")
        return

    # Собираем пары (sol_private_key, proxy, sol_address, account_name) из data
    wallets = []
    for row in rows:
        sol_pk = row.get('sol_private_key', '').strip()
        if not sol_pk:
            continue
        proxy = row.get('proxy', '').strip() or None
        account_name = row.get('name', '').strip() or None
        sol_address = _derive_sol_address(sol_pk)
        if sol_address:
            wallets.append((sol_pk, proxy, sol_address, account_name))

    if not wallets:
        logger.error("Нет SOL приватных ключей в data.csv (поле sol_private_key)")
        return

    # Создаём задачи в БД
    sol_addresses = [w[2] for w in wallets]
    create_check_tasks(sol_addresses)

    # Получаем только незавершённые задачи
    pending = get_pending_tasks()
    pending_addresses = {t['sol_address'] for t in pending}

    # Фильтруем кошельки — только pending
    wallets_to_process = [w for w in wallets if w[2] in pending_addresses]

    if not wallets_to_process:
        pending_total = len(pending)
        if pending_total > 0:
            logger.warning(
                "Для текущего data.csv нет ожидающих задач. "
                "В БД есть pending/failed задачи от других наборов данных."
            )
        else:
            logger.success("Все кошельки из текущего data.csv уже проверены!")
        _print_statistics()
        return

    total = len(wallets_to_process)
    threads = min(NUM_THREADS, total)

    logger.info(f"Начинаем проверку {total} кошельков на Perle eligibility...")
    logger.info(f"Потоков: {threads} │ Попыток: {RETRY_COUNT} │ Задержка: {SLEEP_BETWEEN_ACTIONS}с │ Между аккаунтами: {DELAY_BETWEEN_ACCOUNTS}с")

    # Подготавливаем аргументы для потоков
    tasks = [
        (idx, total, sol_pk, proxy, account_name)
        for idx, (sol_pk, proxy, sol_address, account_name) in enumerate(wallets_to_process, 1)
    ]

    # Мультипоточное выполнение
    with ThreadPoolExecutor(max_workers=threads) as executor:
        futures = {executor.submit(_thread_worker, task): task for task in tasks}

        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                task = futures[future]
                logger.error(f"Ошибка потока [{task[0]:03d}]: {e}")

    # Итоговая статистика
    _print_statistics()

    # Автоэкспорт результатов
    export_path = export_results_xlsx()
    if export_path:
        logger.success(f"Результаты сохранены: {export_path}")


def _print_statistics():
    """Вывод итоговой статистики в консоль."""
    s = get_task_statistics()
    sep = f"{Fore.CYAN}{'═' * 60}{Style.RESET_ALL}"

    lines = [
        "",
        sep,
        f"{Fore.CYAN}  PERLE ELIGIBILITY CHECK — РЕЗУЛЬТАТЫ{Style.RESET_ALL}",
        sep,
        f"  Всего кошельков:   {s['total']}",
        f"  Проверено:         {Fore.GREEN}{s['completed']}{Style.RESET_ALL}",
        f"  Ожидают:           {Fore.YELLOW}{s['pending']}{Style.RESET_ALL}",
        f"  Ошибки:            {Fore.RED}{s['failed']}{Style.RESET_ALL}",
        f"{Fore.CYAN}{'─' * 60}{Style.RESET_ALL}",
        f"  ✅ Eligible:        {Fore.GREEN}{s['eligible']}{Style.RESET_ALL}",
        f"  ❌ Not eligible:    {Fore.RED}{s['not_eligible']}{Style.RESET_ALL}",
        sep,
        "",
    ]
    print("\n".join(lines))


def print_run_statistics():
    """Публичный вызов статистики."""
    _print_statistics()


def export_results_xlsx(output_path: Optional[str] = None) -> Optional[str]:
    """Экспорт результатов в Excel файл"""
    results = get_all_results()
    if not results:
        logger.warning("Нет данных для экспорта")
        return None

    RESULT_DIR.mkdir(parents=True, exist_ok=True)

    if not output_path:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = str(RESULT_DIR / f'perle_eligibility_{timestamp}.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = "Perle Eligibility"

    # Заголовки
    headers = [
        '#', 'SOL Address', 'Status', 'Eligible',
        'Total', 'Details', 'Error', 'Checked At'
    ]

    # Стили
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(name='Consolas', bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Цвета по статусу
    status_colors = {
        'eligible': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
        'not_eligible': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
        'failed': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
        'pending': PatternFill(start_color='D6D8DB', end_color='D6D8DB', fill_type='solid'),
    }

    for idx, row_data in enumerate(results, 1):
        row_num = idx + 1
        is_eligible = bool(row_data.get('eligible', 0))
        eligible_text = 'YES' if is_eligible else 'NO'

        values = [
            idx,
            row_data['sol_address'],
            row_data['status'],
            eligible_text,
            row_data.get('total', '0'),
            row_data.get('details', ''),
            row_data.get('error_message', '') or '',
            row_data.get('completed_at', '') or '',
        ]

        if row_data['status'] == 'completed':
            fill_key = 'eligible' if is_eligible else 'not_eligible'
        elif row_data['status'] == 'failed':
            fill_key = 'failed'
        else:
            fill_key = 'pending'

        row_fill = status_colors[fill_key]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            if col == 2:  # Address
                cell.font = Font(name='Consolas', size=10)
            elif col == 4:  # Eligible
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)

    # Ширина колонок
    col_widths = [5, 48, 12, 10, 12, 40, 40, 22]
    for i, width in enumerate(col_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else 'A'
        ws.column_dimensions[col_letter].width = width

    # Итоговая строка
    total_row = len(results) + 2
    eligible_count = sum(1 for r in results if r.get('eligible', 0))
    not_eligible_count = sum(1 for r in results if r['status'] == 'completed' and not r.get('eligible', 0))

    ws.cell(row=total_row, column=1, value='ИТОГО').font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=f'Всего: {len(results)}').font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=f'Eligible: {eligible_count} / Not: {not_eligible_count}').font = Font(bold=True)

    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).border = thin_border

    try:
        wb.save(output_path)
        logger.success(f"Excel файл сохранён: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Ошибка сохранения Excel: {e}")
        return None
